"""
RFP Backend API - Question Detection and Processing

IMPLEMENTATION NOTE:
===================
This file contains a 1:1 translation from the proven TypeScript implementation
in services/geminiService.ts and services/xlsxService.ts

==============================================================================
COMPLETE FLOW: TypeScript → Python
==============================================================================

1. READ EXCEL (xlsxService.ts: parseXlsx)
   TypeScript: XLSX.utils.sheet_to_json(worksheet, { header: 1 })
   Python:     openpyxl.load_workbook() + _extract_row_data()
   Result:     Array of arrays (rows of cells)

2. PREPARE DATA (TypeScript: create RowWithNumber[])
   TypeScript: rows.map(row => ({rowNumber: idx, rowData: row}))
   Python:     [{"rowNumber": row_num, "rowData": [cells]}, ...]
   Result:     Rows with numbers for AI processing
   VALIDATION: Skip row 1 (header), start from row 2

3. DETECT QUESTIONS (geminiService.ts: detectQuestionsInBatch)
   TypeScript: await detectQuestionsInBatch(rows)
   Python:     detect_questions_in_batch(rows_with_numbers)
   Result:     [{"rowNumber": int, "question": str}, ...]
   VALIDATION: Filter out any invalid row numbers

4. GENERATE ANSWERS (Python-specific - not in TypeScript)
   Python:     process_detected_questions_batch()
   Result:     {rowNumber: {"answer": str, "review_status": str}}
   VALIDATION: Row must be >= 2 and <= max_row

5. PLACE ANSWERS (TypeScript: modify sheet.data)
   TypeScript: sheet.data[rowNumber][colIdx] = answer
   Python:     ws.cell(row=rowNumber, column=ai_col, value=answer)
   Result:     Answers in EXACT detected row numbers
   VALIDATION: Triple-check row is not 1, >= 2, <= max_row

6. SAVE EXCEL (xlsxService.ts: createXlsxAndDownload)
   TypeScript: XLSX.utils.aoa_to_sheet(sheet.data); XLSX.writeFile()
   Python:     wb.save(output)
   Result:     Excel file with answers

==============================================================================

ROW-BY-ROW PROCESSING - DIRECT MAPPING:
==============================================================================

The NEW approach eliminates all searching and ambiguity:

  Excel Row 5: "What is your approach to security?"
       ↓
  Send to AI: {"rowNumber": 5, "rowData": ["What", "is", "your", ...]}
       ↓
  AI Returns: {"rowNumber": 5, "question": "What is your approach..."}
       ↓
  Generate Answer for Row 5: "Our approach includes..."
       ↓
  Place Answer in Row 5: ws.cell(row=5, column=ai_col, value=answer)

✓ Question in Row 5 → Answer in Row 5 (GUARANTEED)
✓ No searching through the sheet
✓ No fuzzy matching
✓ No row number confusion

OLD approach problems (DEPRECATED):
- Sent entire sheet as text
- AI returned questions with uncertain row numbers
- Had to search/match to find correct rows
- Row numbers often mismatched

==============================================================================

ANSWER PLACEMENT RELIABILITY - CRITICAL SAFEGUARDS:
==============================================================================

Problem: Answers must be placed in the SAME ROW as their questions. Never in:
- Row 1 (header row)
- Empty rows
- Rows outside the data range

Solution: Multiple validation layers at each step:

1. INPUT VALIDATION (Row Collection):
   - Start from row 2 (min_data_row = 2)
   - Explicitly check and skip if row < min_data_row
   - Validate no row 1 in collected rows

2. AI PROMPT VALIDATION:
   - Explicitly instruct AI to return EXACT row numbers
   - Warn AI not to modify or recalculate row numbers

3. OUTPUT VALIDATION (After AI Detection):
   - Log all row numbers returned by AI
   - Filter out rows < min_data_row
   - Filter out rows > max_row
   - Remove any header row violations

4. ANSWER GENERATION VALIDATION:
   - Pass min_valid_row and max_valid_row to answer generator
   - Reject any row < min_valid_row
   - Reject any row > max_valid_row

5. PLACEMENT VALIDATION (Triple-Check):
   - Check row != 1 (header)
   - Check row >= min_data_row (2)
   - Check row <= max_row
   - Log each successful placement with question text

6. SUMMARY REPORTING:
   - Show exact rows where answers were placed
   - Show count of questions detected vs placed
   - Show any discrepancies

==============================================================================

Key Functions Mapping:
- TypeScript: detectQuestionsInBatch() → Python: detect_questions_in_batch()
- TypeScript: parseXlsx() → Python: _extract_row_data() + openpyxl
- TypeScript: sheet.data[row] = [..., answer] → Python: ws.cell(row=row, column=col)

See comments starting with "# TypeScript:" for exact mappings throughout the code.
"""

import os
import io
import json
import difflib
import re
import pandas as pd
import openpyxl
from fastapi import FastAPI, UploadFile, Header, HTTPException, Form, File
from fastapi import Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from supabase import create_client, Client
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import google.generativeai as genai
from google.generativeai.types import HarmCategory, HarmBlockThreshold
import zipfile
import tempfile
from dotenv import load_dotenv
import asyncio
import threading
import time
import uuid
import gc
from datetime import datetime, timedelta
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
import traceback # Import traceback module

# CONFIGURATION 
load_dotenv()
def _clean_env(value: str | None) -> str:
    if not value:
        return ""
    return value.strip().strip('"').strip("'")

GOOGLE_API_KEY = _clean_env(os.getenv("GOOGLE_API_KEY"))
SUPABASE_URL = _clean_env(os.getenv("SUPABASE_URL"))
SUPABASE_KEY = _clean_env(os.getenv("SUPABASE_KEY"))

# Initialize Gemini client
if not GOOGLE_API_KEY:
    raise ValueError("Missing GOOGLE_API_KEY for Gemini")
genai.configure(api_key=GOOGLE_API_KEY)
GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-2.5-flash")
gemini = genai.GenerativeModel(GEMINI_MODEL)

# Validate and initialize Supabase client with clearer errors
if not SUPABASE_URL or not SUPABASE_URL.startswith("https://") or ".supabase.co" not in SUPABASE_URL:
    raise ValueError(f"Invalid SUPABASE_URL format: '{SUPABASE_URL}'. Expected like https://xxxxx.supabase.co")
if not SUPABASE_KEY:
    raise ValueError("SUPABASE_KEY is missing")

# Create Supabase client with default configuration
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# FastAPI app
app = FastAPI(
    title="RFP Backend API", 
    version="1.0.0",
    description="Backend API for RFP processing and management",
    docs_url="/docs",
    redoc_url="/redoc"
)

# Root endpoint
@app.get("/")
def root():
    return {
        "message": "RFP Backend API",
        "version": "1.0.0",
        "docs": "/docs",
        "health": "/health"
    }

# Test endpoint to verify connectivity
@app.get("/test")
def test_endpoint():
    print("=== TEST ENDPOINT CALLED ===")
    return {"message": "Backend is working!", "timestamp": datetime.now().isoformat()}

# Health check endpoint for Render monitoring
@app.get("/health")
def health_check():
    """Health check endpoint for monitoring"""
    try:
        # Test database connectivity
        supabase.table("clients").select("id").limit(1).execute()
        return {
            "status": "healthy",
            "timestamp": datetime.now().isoformat(),
            "database": "connected"
        }
    except Exception as e:
        print(f"ERROR: Health check failed: {e}")
        traceback.print_exc()
        return {
            "status": "unhealthy",
            "timestamp": datetime.now().isoformat(),
            "database": "disconnected",
            "error": str(e)
        }

# Test endpoint for job submission debugging
@app.post("/jobs/submit-test")
async def submit_job_test(
    file: UploadFile = File(...),
    job_type: str = Form(...),
    x_client_key: str | None = Header(default=None, alias="X-Client-Key")
):
    """Test endpoint for job submission debugging"""
    print(f"=== /jobs/submit-test ENDPOINT CALLED ===")
    print(f"File: {file.filename}, Job Type: {job_type}, Client Key: {x_client_key}")
    return {"message": "Test endpoint reached", "file": file.filename, "job_type": job_type}

# CORS for frontend dev and configurable origins
frontend_origin = os.getenv("FRONTEND_ORIGIN", "http://localhost:5173")
allowed_origins = [
    frontend_origin, 
    "http://127.0.0.1:5173", 
    "http://localhost:5173",
    "https://localhost:5173",
    "https://127.0.0.1:5173",
    "https://rfp-two.vercel.app"  # Production frontend
]

# Add production origins if they exist
if os.getenv("RENDER"):
    # Running on Render
    allowed_origins.extend([
        "https://rfp-two.vercel.app",  # Production frontend
        "https://*.vercel.app",  # Allow any Vercel subdomain
        "https://*.onrender.com"  # Allow any Render subdomain
    ])

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_origin_regex=r"https?://(localhost|127\\.0\\.0\\.1|.*\\.onrender\\.com|.*\\.vercel\\.app)(:\\d+)?",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=[
        "*",
        "Content-Type",
        "X-Client-Key",
        "X-RFP-ID",
    ],
    expose_headers=[
        "Content-Disposition",
    ],
)

# HELPER FUNCTIONS

def extract_questions_with_gemini(sheet_text: str) -> list:
    # Reverted to user's v1 prompt for question detection
    prompt = f"""
You are analyzing an RFP Excel sheet. Identify ALL rows that represent questions or
requirements directed at the vendor. Treat both explicit questions (with '?') and
implicit requirements (like "Vendor must provide X", "Describe how your system handles Y")
as questions.

Return only a JSON array where each object has:
- "question": the extracted question/requirement text (cleaned and concise).
- "row": the row number in the sheet (1-based).
Return JSON array only:
[
  {{"question": "Question text here", "row": 5}}
]

If no questions are found, return [].

Sheet:
{sheet_text}
"""
    try:
        result = [None]
        error = [None]
        
        def api_call():
            try:
                response = gemini.generate_content(
                    prompt,
                    safety_settings={
                        HarmCategory.HARM_CATEGORY_HARASSMENT: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                        HarmCategory.HARM_CATEGORY_HATE_SPEECH: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                        HarmCategory.HARM_CATEGORY_SEXUALLY_EXPLICIT: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                        HarmCategory.HARM_CATEGORY_DANGEROUS_CONTENT: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                    },
                )
                result[0] = response
            except Exception as e:
                error[0] = e
        
        thread = threading.Thread(target=api_call)
        thread.daemon = True
        thread.start()
        
        thread.join(timeout=60) 
        
        if thread.is_alive():
            print("Gemini extract: API call timed out after 60 seconds")
            return []
        
        if error[0]:
            raise error[0]
        
        if not result[0]:
            print("Gemini extract: No response received")
            return []
        
        text = (result[0].text or "").strip()
        start, end = text.find("["), text.rfind("]")
        if start == -1 or end == -1:
            print(f"Gemini extract: No JSON array found in response for text: {text[:200]}...")
            return []
        return json.loads(text[start:end+1])
            
    except Exception as e:
        print(f"Gemini extract error: {e}")
        traceback.print_exc()
        return []


def _prepare_sheet_context(df, sheet_name: str, ws) -> str:
    """Prepare sheet text with structural context for better AI analysis"""
    
    max_row = ws.max_row
    max_col = ws.max_column
    
    context_parts = [
        f"--- SHEET: '{sheet_name}' ---",
        f"Sheet Dimensions: {max_row} rows x {max_col} columns",
        "",
        "STRUCTURAL CONTEXT (first few rows & columns to identify headers/categories):",
    ]
    
    for r_idx in range(1, min(max_row + 1, 10)): # Check first 10 rows
        row_values = []
        for c_idx in range(1, min(max_col + 1, 15)): # Check first 15 columns
            cell_value = ws.cell(row=r_idx, column=c_idx).value
            if cell_value is not None:
                row_values.append(str(cell_value).strip())
        if row_values:
            context_parts.append(f"Row {r_idx}: {' | '.join(row_values)}")
    
    context_parts.append("")
    context_parts.append("FULL SHEET DATA (to identify specific questions/requirements):")
    
    df_string = df.to_string(index=False, header=False)
    if len(df_string) > 20000: 
        df_string = df_string[:20000] + "\n... [Remaining sheet data truncated]"
    context_parts.append(df_string)
    
    return "\n".join(context_parts)

# Simplified question extraction wrapper
def extract_questions_combined(sheet_text: str) -> list:
    """
    Primary method for question extraction using the simplified Gemini prompt.
    """
    print("DEBUG: Attempting simplified Gemini-based question extraction...")
    gemini_questions = extract_questions_with_gemini(sheet_text)
    if gemini_questions:
        print(f"DEBUG: Gemini extracted {len(gemini_questions)} questions.")
    else:
        print("DEBUG: Simplified Gemini extraction found no questions or failed.")
    
    return gemini_questions


def _extract_row_data(ws, row_num: int) -> list:
    """
    Extract all cell values from a row as a list.
    Equivalent to: XLSX.utils.sheet_to_json(worksheet, { header: 1 })
    Returns array of cell values for a single row.
    """
    max_col = ws.max_column or 1
    row_values = []
    
    for c_idx in range(1, max_col + 1):
        cell_value = ws.cell(row=row_num, column=c_idx).value
        if cell_value is not None:
            row_values.append(str(cell_value).strip())
        else:
            row_values.append("")
    
    return row_values


# ============================================================================
# EXACT TRANSLATION FROM services/geminiService.ts
# ============================================================================
# TypeScript Function: detectQuestionsInBatch(rows: RowWithNumber[])
# Python Function: detect_questions_in_batch(rows_with_numbers: list)
#
# This is a 1:1 translation of the working TypeScript implementation.
# Input:  [{"rowNumber": int, "rowData": [array of cells]}, ...]
# Output: [{"rowNumber": int, "question": str}, ...]
# ============================================================================

def detect_questions_in_batch(rows_with_numbers: list) -> list:
    """
    Detect questions in a batch of rows using the EXACT proven TypeScript approach.
    
    INPUT:  [{"rowNumber": 5, "rowData": ["cell1", "cell2", ...]}, ...]
    OUTPUT: [{"rowNumber": 5, "question": "extracted question text"}, ...]
    
    KEY INSIGHT: We pass row numbers TO the AI, and AI returns the SAME row numbers.
    This eliminates all ambiguity - we know EXACTLY which row contains each question.
    
    NO SEARCHING NEEDED:
    - Old approach: Send whole sheet → AI returns questions → search to find rows
    - New approach: Send rows with numbers → AI returns same row numbers → use directly
    
    This is a 1:1 translation of the TypeScript detectQuestionsInBatch function.
    """
    if not rows_with_numbers or len(rows_with_numbers) == 0:
        return []
    
    # EXACT schema from TypeScript version
    response_schema = {
        "type": "ARRAY",
        "description": "A list of rows that have been identified as questions.",
        "items": {
            "type": "OBJECT",
            "properties": {
                "rowNumber": {
                    "type": "NUMBER",
                    "description": "The original 1-based row number from the input that contains the question."
                },
                "question": {
                    "type": "STRING",
                    "description": "The exact and complete question text extracted from the row."
                }
            },
            "required": ["rowNumber", "question"]
        }
    }
    
    # EXACT prompt from TypeScript version (with proper indentation)
    prompt = f"""
    You are an RFP analysis expert. Your task is to identify which of the following spreadsheet rows are questions or vendor requirements.
    The data is provided as a JSON array, where each object has a 'rowNumber' and 'rowData'.

    Analyze each row's 'rowData'. If it contains a question or a requirement (e.g., "Describe your process...", "Vendor must provide..."), extract it.
    
    Your response MUST be a JSON array of objects, strictly conforming to the provided schema.
    Each object in your response array must correspond to a row that you identified as a question.
    It must include the EXACT ORIGINAL 'rowNumber' and the extracted 'question' text.
    
    CRITICAL RULES:
    1. Return the EXACT 'rowNumber' from the input - DO NOT modify or recalculate row numbers
    2. Only include rows that are actual questions or requirements in your response
    3. Ignore headers, section titles, and empty rows
    4. If no questions are found, return an empty array []
    5. The 'rowNumber' in your response MUST match the 'rowNumber' from the input exactly

    Spreadsheet rows to analyze:
    {json.dumps(rows_with_numbers)}
    """
    
    try:
        result = [None]
        error = [None]
        
        def api_call():
            try:
                # Use gemini-2.5-flash as in TypeScript version
                model = genai.GenerativeModel("gemini-2.5-flash")
                response = model.generate_content(
                    prompt,
                    generation_config={
                        "response_mime_type": "application/json",
                        "response_schema": response_schema
                    },
                    safety_settings={
                        HarmCategory.HARM_CATEGORY_HARASSMENT: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                        HarmCategory.HARM_CATEGORY_HATE_SPEECH: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                        HarmCategory.HARM_CATEGORY_SEXUALLY_EXPLICIT: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                        HarmCategory.HARM_CATEGORY_DANGEROUS_CONTENT: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                    },
                )
                result[0] = response
            except Exception as e:
                error[0] = e
        
        thread = threading.Thread(target=api_call)
        thread.daemon = True
        thread.start()
        
        # 60-second timeout exactly as in TypeScript withTimeout
        thread.join(timeout=60)
        
        if thread.is_alive():
            print("Gemini returned empty response for batch (timeout)")
            return []
        
        if error[0]:
            print(f"Error processing batch with Gemini: {error[0]}")
            traceback.print_exc()
            raise error[0]
        
        if not result[0]:
            print("Gemini returned empty response for batch")
            return []
        
        text = (result[0].text or "").strip()
        if not text:
            print("Gemini returned empty response for batch")
            return []
        
        detected_questions = json.loads(text)
        print(f"DEBUG: Detected {len(detected_questions)} questions in batch")
        return detected_questions
        
    except Exception as e:
        print(f"Error processing batch with Gemini: {e}")
        traceback.print_exc()
        return []


def process_detected_questions_batch(detected_questions: list, client_id: str, rfp_id: str, min_valid_row: int = 2, max_valid_row: int = None) -> dict:
    """
    Process detected questions to generate answers for SPECIFIC rows.
    
    INPUT:  [{"rowNumber": 5, "question": "What is your approach?"}, ...]
    OUTPUT: {5: {"question": "What...", "answer": "Our approach...", "review_status": "..."}}
    
    The returned dictionary maps row numbers directly - no searching needed!
    Row 5 question → Row 5 answer (exact match)
    
    VALIDATION: Ensures answers are never placed in:
    - Row 1 (header row)
    - Rows outside the valid data range (< min_valid_row or > max_valid_row)
    """
    results = {}
    
    for item in detected_questions:
        row_num = item.get("rowNumber", 0)
        question_text = item.get("question", "").strip()
        
        # CRITICAL VALIDATION: Ensure row number is valid
        if not question_text or row_num == 0:
            print(f"WARN: Skipping item with empty question or row 0")
            continue
        
        # NEVER place answers in row 1 (header row)
        if row_num < min_valid_row:
            print(f"ERROR: AI returned invalid row number {row_num} (below min {min_valid_row}). Skipping to prevent header corruption.")
            continue
        
        # Check if row exceeds max valid row
        if max_valid_row and row_num > max_valid_row:
            print(f"ERROR: AI returned invalid row number {row_num} (above max {max_valid_row}). Skipping.")
            continue
        
        try:
            # Search Supabase for answer
            emb = get_embedding(question_text)
            matches = search_supabase(emb, client_id, rfp_id)
            
            final_answer = "Not found, needs review."
            review_status = "Need review"
            
            if matches:
                best_match = pick_best_match(matches)
                if best_match and best_match.get("similarity", 0) >= 0.9:
                    final_answer = best_match["answer"]
                    review_status = "Approved"
                else:
                    # Use tailored answer from matches
                    final_answer = generate_tailored_answer(question_text, matches)
                    review_status = "Need review - AI Generated"
            
            results[row_num] = {
                "question": question_text,
                "answer": clean_markdown(final_answer),
                "review_status": review_status
            }
            print(f"DEBUG: Generated answer for row {row_num}: '{question_text[:50]}...'")
            
        except Exception as e:
            print(f"ERROR: Error processing question for row {row_num}: {e}")
            traceback.print_exc()
            continue
    
    return results


def clean_markdown(text: str) -> str:
    """Remove common Markdown formatting such as **bold**, _italic_, `code`, headings, and links.
    Keeps only readable plain text with proper wrapping.
    """
    if not text:
        return ""
    cleaned = text
    
    cleaned = re.sub(r"!\[[^\]]*\]\([^)]*\)", "", cleaned)
    cleaned = re.sub(r"\[([^\]]+)\]\([^)]*\)", r"\1", cleaned)
    cleaned = re.sub(r"`([^`]*)`", r"\1", cleaned)
    
    cleaned = re.sub(r"\*\*([^*]+)\*\*", r"\1", cleaned)
    cleaned = re.sub(r"__([^_]+)__", r"\1", cleaned)
    cleaned = re.sub(r"\*([^*]+)\*", r"\1", cleaned)
    cleaned = re.sub(r"_([^_]+)_", r"\1", cleaned)
    
    cleaned = re.sub(r"^\s{0,3}#{1,6}\s*", "", cleaned, flags=re.MULTILINE)
    cleaned = re.sub(r"^\s{0,3}>\s?", "", cleaned, flags=re.MULTILINE)
    cleaned = re.sub(r"^\s*[-*+]\s+", "", cleaned, flags=re.MULTILINE)
    cleaned = re.sub(r"^\s*\d+\.\s+", "", cleaned, flags=re.MULTILINE)
    cleaned = re.sub(r"^\s*(?:-{3,}|\*{3,}|_{3,})\s*$", "", cleaned, flags=re.MULTILINE)
    
    cleaned = re.sub(r"\s+\n", "\n", cleaned)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    cleaned = re.sub(r'\s*([.,!?;:])', r'\1', cleaned)
    cleaned = re.sub(r'^\s+|\s+$', '', cleaned, flags=re.MULTILINE)
    cleaned = re.sub(r' {2,}', ' ', cleaned)
    
    return cleaned.strip()


def get_embedding(text: str) -> list:
    if not text:
        print("DEBUG: Empty text provided to get_embedding")
        return []
    try:
        res = genai.embed_content(model="models/embedding-001",
                                  content=text,
                                  task_type="retrieval_query")
        embedding = res["embedding"]
        return embedding
    except Exception as e:
        print(f"Embedding error: {e} for text: {text[:100]}...")
        traceback.print_exc()
        return []


def search_supabase(question_embedding: list, client_id: str, rfp_id: str = None) -> list:
    if not question_embedding:
        print("DEBUG: No embedding provided to search_supabase")
        return []
    try:
        print(f"DEBUG: Searching Supabase with client_id={client_id}, rfp_id={rfp_id}")
        # Using 'client_match_questions' RPC as it's designed for the current multi-client setup
        res = supabase.rpc(
            "client_match_questions", {
                "query_embedding": question_embedding,
                "match_threshold": 0.0, # Keep 0.0 here for broader search, filter later by 0.9 for best match
                "match_count": 5,
                "p_client_id": client_id,
                "p_rfp_id": rfp_id
            }).execute()
        print(f"DEBUG: Supabase RPC response: {res.data if res.data else 'None'}")
        return res.data if res.data else []
    except Exception as e:
        print(f"Supabase search error: {e}")
        traceback.print_exc()
        return []


def pick_best_match(matches: list):
    if not matches:
        return None
    return max(matches, key=lambda m: m.get("similarity", 0))

def generate_tailored_answer(question: str, matches: list) -> str:
    context = "\n".join(
        f"- Q: {m['question']} | A: {m['answer']} (similarity {m['similarity']:.2f})"
        for m in matches)
    prompt = f"""
You are answering an RFP vendor question.  

New Question:
{question}

Reference Q&A Pairs:
{context}

Write a concise, tailored answer that best addresses the new question using the references.
If unclear, combine and adapt from references.
Return only the answer text, without any additional conversational filler or markdown formatting.
"""
    try:
        result = [None]
        error = [None]
        
        def api_call():
            try:
                resp = gemini.generate_content(
                    prompt,
                    safety_settings={
                        HarmCategory.HARM_CATEGORY_HARASSMENT: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                        HarmCategory.HARM_CATEGORY_HATE_SPEECH: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                        HarmCategory.HARM_CATEGORY_SEXUALLY_EXPLICIT: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                        HarmCategory.HARM_CATEGORY_DANGEROUS_CONTENT: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                    },
                )
                result[0] = resp
            except Exception as e:
                error[0] = e
        
        thread = threading.Thread(target=api_call)
        thread.daemon = True
        thread.start()
        
        thread.join(timeout=45)
        
        if thread.is_alive():
            print("Gemini tailored answer: API call timed out after 45 seconds")
            return "AI could not generate tailored answer (timeout)."
        
        if error[0]:
            raise error[0]
        
        if not result[0]:
            print("Gemini tailored answer: No response received")
            return "AI could not generate tailored answer."
        
        return result[0].text.strip()
            
    except Exception as e:
        print(f"Gemini tailored answer error: {e} for question: {question[:100]}...")
        traceback.print_exc()
        return "AI could not generate tailored answer."


# ==============================================================================
# DEPRECATED: Old fuzzy matching functions - NO LONGER USED
# ==============================================================================
# The new row-by-row processing approach directly uses the row number from AI
# detection. No fuzzy matching or searching is needed because:
# 1. AI receives: {"rowNumber": 5, "rowData": ["cell1", "cell2", ...]}
# 2. AI returns: {"rowNumber": 5, "question": "extracted question"}
# 3. We place answer directly in row 5
#
# These functions are kept for reference but are NOT called in the main flow.
# ==============================================================================

def _row_text(ws, r: int) -> str:
    """DEPRECATED: Extracts concatenated text from a given row, ignoring empty cells."""
    return " ".join(
        str(ws.cell(row=r, column=c).value or "").strip().lower()
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=r, column=c).value is not None and str(ws.cell(row=r, column=c).value).strip() != ''
    )

def resolve_row(worksheet, reported_row: int, question_text: str) -> int:
    """DEPRECATED: Old fuzzy matching approach - NO LONGER USED
    
    With the new row-by-row processing:
    - AI tells us the exact row number where it found the question
    - We use that row number directly
    - No searching or matching needed
    """
    max_row = worksheet.max_row
    min_data_row = 2  # Never write answers to headers, always start from row 2
    
    # Prioritize reported row if valid and has content
    if reported_row and min_data_row <= reported_row <= max_row:
        row_content_check = _row_text(worksheet, reported_row)
        if row_content_check: # Check if the row actually has content
            print(f"DEBUG: Resolved row for '{question_text[:50]}...' using reported row {reported_row} (contains content).")
            return reported_row
        else:
            print(f"DEBUG: Reported row {reported_row} for '{question_text[:50]}...' is empty, falling back to fuzzy search.")
    
    qnorm = (question_text or "").strip().lower()
    best_r, best_score = None, 0.0 # Use 0.0 for float comparison

    # Fuzzy search from min_data_row onwards
    for r in range(min_data_row, max_row + 1):
        row_text = " ".join(
            str(worksheet.cell(row=r, column=c).value or "").lower()
            for c in range(1, worksheet.max_column + 1))
        
        # Skip empty or very short rows in fuzzy matching, similar to v1 implicit behavior
        if not row_text.strip() or len(row_text.strip()) < 10: 
            continue
            
        s = difflib.SequenceMatcher(None, qnorm, row_text).ratio()
        if s > best_score: # Pick the best match, no hard threshold (v1 behavior)
            best_score, best_r = s, r
    
    # Ensure best_r is not None and is a valid data row, else default to min_data_row
    resolved_row = best_r if best_r is not None and best_r >= min_data_row else min_data_row
    print(f"DEBUG: Resolved row for '{question_text[:50]}...' using fuzzy search -> {resolved_row} (Score: {best_score:.2f})")
    return resolved_row


def find_first_empty_data_column(ws):
    """
    Finds the first entirely empty column after the rightmost column that contains
    a header in the first row OR any data in any subsequent row.
    Ensures it's not column A (index 1).
    """
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1 
    
    rightmost_filled_col = 1
    # Check row 1 for headers
    for col in range(1, max_col + 1):
        header_cell_value = ws.cell(row=1, column=col).value
        if header_cell_value is not None and str(header_cell_value).strip() != '':
            rightmost_filled_col = max(rightmost_filled_col, col)

    # Check data rows (from row 2 onwards) for content
    for col in range(1, max_col + 1):
        for row in range(2, max_row + 1): 
            data_cell_value = ws.cell(row=row, column=col).value
            if data_cell_value is not None and str(data_cell_value).strip() != '':
                rightmost_filled_col = max(rightmost_filled_col, col)
                break 

    # Start search for empty column from the column immediately after the rightmost filled column.
    # Explicitly ensure AI answers don't go into column A.
    candidate_start_col = max(rightmost_filled_col + 1, 2) 

    current_candidate_col = candidate_start_col
    
    # Add a safety break to prevent infinite loops on malformed or extremely dense sheets
    search_limit = max_col + 50 # Search up to 50 columns beyond current max_col
    while current_candidate_col <= search_limit:
        is_column_empty = True
        for r in range(1, max_row + 1): # Check all rows in the candidate column
            if ws.cell(row=r, column=current_candidate_col).value is not None and \
               str(ws.cell(row=r, column=current_candidate_col).value).strip() != '':
                is_column_empty = False
                break
        
        if is_column_empty:
            print(f"DEBUG: Identified column {get_column_letter(current_candidate_col)} as the first empty data column.")
            return current_candidate_col
        
        current_candidate_col += 1
        
    print(f"WARN: Could not find an entirely empty column after checking {search_limit} columns. Returning {current_candidate_col-1} as fallback.")
    return current_candidate_col -1 # Return the last checked column + 1 as a fallback, or if no empty found, the next

# API ENDPOINT
def get_client_id_from_key(client_key: str | None) -> str:
    global supabase
    if not client_key:
        raise HTTPException(status_code=401, detail="Missing X-Client-Key")
    max_retries = 3
    for attempt in range(max_retries):
        try:
            resp = supabase.table("clients").select("id").eq("api_key", client_key).limit(1).execute()
            rows = resp.data or []
            if not rows:
                raise HTTPException(status_code=401, detail="Invalid X-Client-Key")
            return rows[0]["id"]
        except HTTPException:
            raise
        except Exception as e:
            print(f"Client lookup error: {e}")
            traceback.print_exc()
            # Recreate supabase client and retry on transient errors
            try:
                if supabase is None:
                    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
                else:
                    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
            except Exception as reinit_err:
                print(f"Supabase re-init failed: {reinit_err}")
                traceback.print_exc()
            if attempt < max_retries - 1:
                try:
                    time.sleep(0.5)
                except Exception:
                    pass
                continue
        raise HTTPException(status_code=500, detail="Client lookup failed")


@app.post("/process")
async def process_excel(file: UploadFile, x_client_key: str | None = Header(default=None, alias="X-Client-Key"), rfp_id: str | None = Header(default=None, alias="X-RFP-ID")):
    client_id = get_client_id_from_key(x_client_key)
    
    file_content = await file.read()
    file_obj = io.BytesIO(file_content)
    
    processed_file_io, processed_sheets_count, total_questions_processed = process_excel_file_obj(file_obj, file.filename, client_id, rfp_id)
    
    return StreamingResponse(
        processed_file_io,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=processed_{file.filename}"}
    )


# --- RFP management routes ---

@app.get("/rfps")
def list_rfps(x_client_key: str | None = Header(default=None, alias="X-Client-Key")):
    print(f"=== /rfps ENDPOINT CALLED ===")
    print("RFPs endpoint called")
    client_id = get_client_id_from_key(x_client_key)
    print(f"Resolved client_id: {client_id}")
    try:
        res = supabase.table("client_rfps").select("id, name, description, created_at, updated_at").eq("client_id", client_id).order("created_at", desc=True).execute()
        print(f"Found {len(res.data or [])} RFPs")
        return {"rfps": res.data or []}
    except Exception as e:
        print(f"Error listing RFPs: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Failed to list RFPs")

@app.post("/rfps")
def create_rfp(payload: dict = Body(...), x_client_key: str | None = Header(default=None, alias="X-Client-Key")):
    client_id = get_client_id_from_key(x_client_key)
    rfp_data = {
        "client_id": client_id,
        "name": payload.get("name", "").strip(),
        "description": payload.get("description", "").strip(),
    }
    if not rfp_data["name"]:
        raise HTTPException(status_code=400, detail="RFP name is required")
    try:
        res = supabase.table("client_rfps").insert(rfp_data).execute()
        return {"rfp": res.data[0] if res.data else None}
    except Exception as e:
        print(f"Error creating RFP: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Failed to create RFP")

@app.put("/rfps/{rfp_id}")
def update_rfp(rfp_id: str, payload: dict = Body(...), x_client_key: str | None = Header(default=None, alias="X-Client-Key")):
    client_id = get_client_id_from_key(x_client_key)
    updates = {k: v for k, v in payload.items() if k in ("name", "description")}
    if "name" in updates:
        updates["name"] = updates["name"].strip()
    if "description" in updates:
        updates["description"] = updates["description"].strip()
    try:
        res = supabase.table("client_rfps").update(updates).eq("id", rfp_id).eq("client_id", client_id).execute()
        return {"ok": True}
    except Exception as e:
        print(f"Error updating RFP {rfp_id}: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Failed to update RFP")

@app.delete("/rfps/{rfp_id}")
def delete_rfp(rfp_id: str, x_client_key: str | None = Header(default=None, alias="X-Client-Key")):
    client_id = get_client_id_from_key(x_client_key)
    try:
        supabase.table("client_rfps").delete().eq("id", rfp_id).eq("client_id", client_id).execute()
        return {"ok": True}
    except Exception as e:
        print(f"Error deleting RFP {rfp_id}: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Failed to delete RFP")

# --- Organization-focused routes ---

@app.get("/org")
def get_org(x_client_key: str | None = Header(default=None, alias="X-Client-Key")):
    client_id = get_client_id_from_key(x_client_key)
    try:
        res = supabase.table("clients").select("id, name, sector, contact_email").eq("id", client_id).single().execute()
        return res.data
    except Exception as e:
        print(f"Error getting organization data: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Failed to get organization data")


@app.put("/org")
def update_org(payload: dict = Body(...), x_client_key: str | None = Header(default=None, alias="X-Client-Key")):
    client_id = get_client_id_from_key(x_client_key)
    updates = {k: v for k, v in payload.items() if k in ("name", "sector", "contact_email")}
    try:
        supabase.table("clients").update(updates).eq("id", client_id).execute()
        return {"ok": True}
    except Exception as e:
        print(f"Error updating organization data: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Failed to update organization data")


@app.get("/org/qa")
def list_org_qa(x_client_key: str | None = Header(default=None, alias="X-Client-Key"), rfp_id: str | None = Header(default=None, alias="X-RFP-ID")):
    try:
        client_id = get_client_id_from_key(x_client_key)
        qs_query = supabase.table("client_questions").select("id, original_text, category, created_at, rfp_id").eq("client_id", client_id)
        ans_query = supabase.table("client_answers").select("id, answer_text, quality_score, last_updated, rfp_id").eq("client_id", client_id)

        if rfp_id:
            qs_query = qs_query.eq("rfp_id", rfp_id)
            ans_query = ans_query.eq("rfp_id", rfp_id)

        qs = qs_query.order("created_at", desc=True).execute().data or []
        ans = ans_query.order("last_updated", desc=True).execute().data or []

        q_ids = [q.get("id") for q in qs] if qs else []
        mappings = []
        if q_ids:
            mappings = supabase.table("client_question_answer_mappings").select("question_id, answer_id").in_("question_id", q_ids).execute().data or []

        return {"questions": qs, "answers": ans, "mappings": mappings}
    except HTTPException:
        raise
    except Exception as e:
        print(f"list_org_qa error: {e}")
        traceback.print_exc()
        return {"questions": [], "answers": [], "mappings": []}


@app.post("/org/qa")
def ingest_org_qa(payload: dict = Body(...), x_client_key: str | None = Header(default=None, alias="X-Client-Key"), rfp_id: str | None = Header(default=None, alias="X-RFP-ID")):
    """Ingest a list of question/answer pairs for the organization.
    payload: { pairs: [{question, answer, category?}] }
    """
    client_id = get_client_id_from_key(x_client_key)
    pairs = payload.get("pairs", []) or []
    created = 0
    for p in pairs:
        qtext = (p.get("question") or "").strip()
        atext = (p.get("answer") or "").strip()
        category = (p.get("category") or "Other").strip() or "Other"
        if not qtext or not atext:
            continue
        try:
            q_emb = get_embedding(qtext)
            q_row = {
                "original_text": qtext,
                "normalized_text": qtext.lower(),
                "embedding": q_emb,
                "category": category,
                "client_id": client_id,
                "rfp_id": rfp_id,
            }
            q_ins = supabase.table("client_questions").insert(q_row).execute()
            q_id = (q_ins.data or [{}])[0].get("id")

            a_row = {
                "answer_text": atext,
                "answer_type": "General",
                "character_count": len(atext),
                "technical_level": 1,
                "client_id": client_id,
                "rfp_id": rfp_id,
            }
            a_ins = supabase.table("client_answers").insert(a_row).execute()
            a_id = (a_ins.data or [{}])[0].get("id")

            if q_id and a_id:
                supabase.table("client_question_answer_mappings").insert({
                    "question_id": q_id,
                    "answer_id": a_id,
                    "confidence_score": 1.0,
                    "context_requirements": None,
                    "stakeholder_approved": False,
                }).execute()
                created += 1
        except Exception as e:
            print(f"ingest_org_qa error: {e}")
            traceback.print_exc()
            continue
    return {"created": created}


# --- Q/A extraction from uploaded RFPs (ZIP or XLSX) ---

def _extract_qa_pairs(sheet_content_csv: str) -> list:
    prompt = f"""
Analyze the following Excel sheet data and identify question-answer pairs.

- Detect rows where a question and its corresponding answer are present.
- Return them as a JSON array of objects in this exact format:
  [
    {{
      "question": "string",
      "answer": "string",
      "row": 0,
      "category": "Wifi | Map | Other"
    }}
  ]

Rules:
1. Do not include any explanation or extra text outside the JSON array.
2. If you are unsure about a category, default to "Other".
3. If no valid pairs are found, return an empty array [].

Sheet Data:
{sheet_content_csv}
"""
    try:
        response = gemini.generate_content(prompt)
        text = (response.text or "").strip()
        start, end = text.find("["), text.rfind("]")
        if start == -1 or end == -1:
            return []
        pairs = json.loads(text[start:end+1])
        # If answer exists but question missing, synthesize a fallback question
        normalized = []
        for p in pairs:
            q = (p.get("question") or "").strip()
            a = (p.get("answer") or "").strip()
            if not a:
                continue
            if not q:
                q = "Auto-generated question from answer"
            normalized.append({
                "question": q,
                "answer": a,
                "row": p.get("row", 0),
                "category": (p.get("category") or "Other").strip() or "Other",
            })
        return normalized
    except Exception as e:
        print(f"extract_qa_pairs error: {e}")
        traceback.print_exc()
        return []


def _insert_qa_pair(client_id: str, question_text: str, answer_text: str, category: str = "Other", rfp_id: str = None) -> bool:
    try:
        q_emb = get_embedding(question_text)
        q_ins = supabase.table("client_questions").insert({
            "original_text": question_text,
            "normalized_text": (question_text or "").lower(),
            "embedding": q_emb,
            "category": category or "Other",
            "client_id": client_id,
            "rfp_id": rfp_id,
        }).execute()
        q_id = (q_ins.data or [{}])[0].get("id")

        a_ins = supabase.table("client_answers").insert({
            "answer_text": answer_text,
            "answer_type": "General",
            "character_count": len(answer_text or ""),
            "technical_level": 1,
            "client_id": client_id,
            "rfp_id": rfp_id,
        }).execute()
        a_id = (a_ins.data or [{}])[0].get("id")

        if q_id and a_id:
            supabase.table("client_question_answer_mappings").insert({
                "question_id": q_id,
                "answer_id": a_id,
                "confidence_score": 1.0,
                "context_requirements": None,
                "stakeholder_approved": False,
            }).execute()
            return True
    except Exception as e:
        print(f"_insert_qa_pair error: {e}")
        traceback.print_exc()
    return False


@app.post("/org/qa/extract")
async def extract_qa_from_upload(file: UploadFile, x_client_key: str | None = Header(default=None, alias="X-Client-Key"), rfp_id: str | None = Header(default=None, alias="X-RFP-ID")):
    client_id = get_client_id_from_key(x_client_key)
    created = 0
    with tempfile.TemporaryDirectory() as td:
        tmp_path = os.path.join(td, file.filename)
        content = await file.read()
        with open(tmp_path, "wb") as f:
            f.write(content)

        def _process_excel(path: str):
            nonlocal created
            try:
                xls = pd.ExcelFile(path)
                for sheet in xls.sheet_names:
                    df = pd.read_excel(path, sheet_name=sheet, header=None)
                    if df.empty:
                        continue
                    sheet_csv = df.to_csv(index=False, header=False)
                    pairs = _extract_qa_pairs(sheet_csv) or []
                    for p in pairs:
                        q = (p.get("question") or "").strip()
                        a = (p.get("answer") or "").strip()
                        c = (p.get("category") or "Other").strip() or "Other"
                        if q and a and _insert_qa_pair(client_id, q, a, c, rfp_id):
                            created += 1
            except Exception as e:
                print(f"extract_qa_from_upload excel error: {e}")
                traceback.print_exc()

        if file.filename.lower().endswith(".zip"):
            try:
                extract_dir = os.path.join(td, "unzipped")
                os.makedirs(extract_dir, exist_ok=True)
                with zipfile.ZipFile(tmp_path, 'r') as zip_ref:
                    zip_ref.extractall(extract_dir)
                for root, _, files in os.walk(extract_dir):
                    for fn in files:
                        if fn.lower().endswith(".xlsx"):
                            _process_excel(os.path.join(root, fn))
            except Exception as e:
                print(f"extract_qa_from_upload zip error: {e}")
                traceback.print_exc()
        elif file.filename.lower().endswith(".xlsx"):
            _process_excel(tmp_path)
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type. Upload .zip or .xlsx")

    return {"created": created}


# --- Quality scoring trigger ---

def _build_scoring_prompt(question: str, answer: str, reference: str) -> str:
    return f"""
You are an expert RFP evaluator. Given QUESTION, ANSWER, and REFERENCE, return a strict JSON object:
{{ "score": <integer 0-100>, "notes": "<one sentence>" }}

QUESTION:
\"\"\"{question}\"\"\"

ANSWER:
\"\"\"{answer}\"\"\"

REFERENCE:
\"\"\"{reference}\"\"\"
"""


def _extract_score(resp_text: str) -> int:
    if not resp_text:
        return -1
    try:
        s, e = resp_text.find("{"), resp_text.rfind("}")
        if s != -1 and e != -1 and e > s:
            data = json.loads(resp_text[s:e+1])
            if "score" in data:
                val = int(round(float(data["score"])) )
                return max(0, min(100, val))
    except Exception:
        pass
    # fallback: first integer
    m = re.search(r"(\d{1,3})(?:\.\d+)?", resp_text)
    if m:
        try:
            return max(0, min(100, int(round(float(m.group(1))))))
        except Exception:
            return -1
    return -1


@app.post("/org/qa/analyze-similarities")
def analyze_qa_similarities(payload: dict = Body(default={}), x_client_key: str | None = Header(default=None, alias="X-Client-Key"), rfp_id: str | None = Header(default=None, alias="X-RFP-ID")):
    """Analyze QA pairs to find similar questions and create summary suggestions"""
    client_id = get_client_id_from_key(x_client_key)
    similarity_threshold = payload.get("similarity_threshold", 0.85)  # 85% similarity by default
    
    try:
        # Get all QA pairs with embeddings
        questions_res = supabase.table("client_questions").select("id, original_text, embedding, rfp_id").eq("client_id", client_id)
        if rfp_id:
            questions_res = questions_res.eq("rfp_id", rfp_id)
        questions = questions_res.execute().data or []
        
        if len(questions) < 2:
            return {"groups": [], "message": "Not enough QA pairs to analyze"}
        
        # Get answer mappings
        q_ids = [q["id"] for q in questions]
        mappings_res = supabase.table("client_question_answer_mappings").select("question_id, answer_id").in_("question_id", q_ids).execute()
        mappings = mappings_res.data or []
        mapping_dict = {m["question_id"]: m["answer_id"] for m in mappings}
        
        # Get answers
        a_ids = list(set(mapping_dict.values()))
        if not a_ids:
            return {"groups": [], "message": "No answers found for questions"}
        
        answers_res = supabase.table("client_answers").select("id, answer_text").in_("id", a_ids).execute()
        answers = answers_res.data or []
        answer_dict = {a["id"]: a["answer_text"] for a in answers}
        
        # Find similar question groups using embeddings
        similar_groups = []
        processed_questions = set()
        
        for i, q1 in enumerate(questions):
            if q1["id"] in processed_questions:
                continue
                
            emb1 = q1.get("embedding")
            if not emb1 or not isinstance(emb1, list):
                continue
            
            group = [q1]
            processed_questions.add(q1["id"])
            
            # Find similar questions
            for j, q2 in enumerate(questions):
                if i >= j or q2["id"] in processed_questions:
                    continue
                
                emb2 = q2.get("embedding")
                if not emb2 or not isinstance(emb2, list):
                    continue
                
                # Calculate cosine similarity
                try:
                    import numpy as np
                    similarity = np.dot(emb1, emb2) / (np.linalg.norm(emb1) * np.linalg.norm(emb2))
                    
                    if similarity >= similarity_threshold:
                        group.append(q2)
                        processed_questions.add(q2["id"])
                except Exception as e:
                    print(f"Error calculating similarity: {e}")
                    traceback.print_exc()
                    continue
            
            # Only keep groups with 2+ questions
            if len(group) >= 2:
                # Create summary using AI
                questions_text = "\n".join([f"- {q['original_text']}" for q in group])
                answers_text = "\n".join([f"- {answer_dict.get(mapping_dict.get(q['id']), 'No answer')}" for q in group])
                
                summary_prompt = f"""
You are consolidating similar Q&A pairs into a single comprehensive Q&A pair.

Similar Questions:
{questions_text}

Their Answers:
{answers_text}

Create:
1. A consolidated question that covers all the similar questions
2. A comprehensive answer that combines all the answers

Return ONLY a JSON object with this exact format:
{{"question": "consolidated question here", "answer": "comprehensive answer here"}}
"""
                
                try:
                    response = gemini.generate_content(
                        summary_prompt,
                        safety_settings={
                            HarmCategory.HARM_CATEGORY_HARASSMENT: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                            HarmCategory.HARM_CATEGORY_HATE_SPEECH: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                            HarmCategory.HARM_CATEGORY_SEXUALLY_EXPLICIT: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                            HarmCategory.HARM_CATEGORY_DANGEROUS_CONTENT: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                        },
                    )
                    
                    resp_text = (response.text or "").strip()
                    start, end = resp_text.find("{"), resp_text.rfind("}")
                    if start != -1 and end != -1:
                        summary_data = json.loads(resp_text[start:end+1])
                        
                        similar_groups.append({
                            "group_id": f"group_{len(similar_groups)}",
                            "questions": [
                                {
                                    "id": q["id"],
                                    "text": q["original_text"],
                                    "answer": answer_dict.get(mapping_dict.get(q["id"]), "No answer")
                                } for q in group
                            ],
                            "suggested_question": summary_data.get("question", ""),
                            "suggested_answer": summary_data.get("answer", ""),
                            "similarity_count": len(group)
                        })
                except Exception as e:
                    print(f"Error creating summary: {e}")
                    traceback.print_exc()
                    continue
        
        return {
            "groups": similar_groups,
            "total_groups": len(similar_groups),
            "total_questions_analyzed": len(questions),
            "similarity_threshold": similarity_threshold
        }
        
    except Exception as e:
        print(f"Error analyzing similarities: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Analysis failed: {str(e)}")


@app.post("/org/qa/ai-group")
def ai_group_qa(payload: dict = Body(default={}), x_client_key: str | None = Header(default=None, alias="X-Client-Key"), rfp_id: str | None = Header(default=None, alias="X-RFP-ID")):
    """AI-driven grouping of all Q&A pairs for a client (optionally scoped to an RFP).
    Returns groups with question_ids and consolidated Q/A without using embeddings.
    Response schema:
    { "groups": [ { "question_ids": [..], "consolidated_question": str, "consolidated_answer": str } ] }
    """
    client_id = get_client_id_from_key(x_client_key)
    try:
        # Fetch all questions
        q_query = supabase.table("client_questions").select("id, original_text, rfp_id").eq("client_id", client_id)
        if rfp_id:
            q_query = q_query.eq("rfp_id", rfp_id)
        questions = q_query.order("created_at", desc=True).execute().data or []

        if not questions:
            return {"groups": [], "message": "No questions found"}

        # Fetch mappings for answers
        q_ids = [q["id"] for q in questions]
        m_rows = supabase.table("client_question_answer_mappings").select("question_id, answer_id").in_("question_id", q_ids).execute().data or []
        q_to_a = {m["question_id"]: m["answer_id"] for m in m_rows}

        # Fetch answers
        a_ids = list({aid for aid in q_to_a.values() if aid})
        a_rows = []
        if a_ids:
            a_rows = supabase.table("client_answers").select("id, answer_text").in_("id", a_ids).execute().data or []
        a_map = {a["id"]: a.get("answer_text", "") for a in a_rows}

        # Build compact input for the LLM
        qa_lines = []
        for q in questions:
            qid = q["id"]
            qtext = (q.get("original_text") or "").strip()
            atext = (a_map.get(q_to_a.get(qid)) or "").strip()
            if not qtext:
                continue
            qa_lines.append({"id": qid, "q": qtext, "a": atext})

        if not qa_lines:
            return {"groups": [], "message": "No Q&A pairs available"}

        # Prompt Gemini to group and summarize
        qa_text = "\n".join([f"ID:{row['id']}\nQ:{row['q']}\nA:{row['a']}" for row in qa_lines])
        prompt = f"""
You will receive a list of Q&A pairs for a single client (and optionally a specific RFP). Group semantically similar questions together and produce a consolidated Q&A for each group.

Return ONLY strict JSON with this structure:
{{
  "groups": [
    {{
      "question_ids": ["<id>", "<id>", ...],
      "consolidated_question": "string",
      "consolidated_answer": "string"
    }}
  ]
}}

Rules:
1) "question_ids" must be the exact IDs provided.
2) Every ID used must exist in the input. Do not invent IDs.
3) Do not include any text before or after the JSON.

Q&A LIST:
{qa_text}
"""
        try:
            response = gemini.generate_content(
                prompt,
                safety_settings={
                    HarmCategory.HARM_CATEGORY_HARASSMENT: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                    HarmCategory.HARM_CATEGORY_HATE_SPEECH: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                    HarmCategory.HARM_CATEGORY_SEXUALLY_EXPLICIT: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                    HarmCategory.HARMS_CATEGORY_DANGEROUS_CONTENT: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                },
            )
            text = (response.text or "").strip()
            s, e = text.find("{"), text.rfind("}")
            if s == -1 or e == -1:
                return {"groups": [], "message": "LLM returned no JSON"}
            data = json.loads(text[s:e+1])

            # Normalize output
            raw_groups = data.get("groups") or []
            valid_ids_local = {str(x["id"]) for x in questions}
            groups = []
            for g in raw_groups:
                qids = [str(x) for x in (g.get("question_ids") or []) if str(x) in valid_ids_local]
                if not qids:
                    continue
                groups.append({
                    "question_ids": qids,
                    "consolidated_question": (g.get("consolidated_question") or "").strip(),
                    "consolidated_answer": (g.get("consolidated_answer") or "").strip(),
                })
            return {"groups": groups}
        except Exception as e:
            print(f"ai_group_qa LLM error: {e}")
            traceback.print_exc()
            return {"groups": [], "message": "AI grouping failed"}

    except HTTPException:
        raise
    except Exception as e:
        print(f"ai_group_qa error: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Failed to group Q&A")


@app.post("/org/qa/approve-summary")
def approve_qa_summary(payload: dict = Body(...), x_client_key: str | None = Header(default=None, alias="X-Client-Key"), rfp_id: str | None = Header(default=None, alias="X-RFP-ID")):
    """Approve a summary - adds consolidated QA plus summary records and mappings.
    This preserves original questions; it no longer deletes them.
    """
    client_id = get_client_id_from_key(x_client_key)
    
    question_ids = payload.get("question_ids", [])
    consolidated_question = payload.get("consolidated_question", "").strip()
    consolidated_answer = payload.get("consolidated_answer", "").strip()
    
    if not question_ids or not consolidated_question or not consolidated_answer:
        raise HTTPException(status_code=400, detail="Missing required fields")
    
    try:
        # Create embedding for consolidated question
        q_emb = get_embedding(consolidated_question)

        # Insert consolidated question
        q_row = {
            "original_text": consolidated_question,
            "normalized_text": consolidated_question.lower(),
            "embedding": q_emb,
            "category": "Consolidated",
            "client_id": client_id,
            "rfp_id": rfp_id,
        }
        q_ins = supabase.table("client_questions").insert(q_row).execute()
        new_q_id = (q_ins.data or [{}])[0].get("id")

        # Insert consolidated answer
        a_row = {
            "answer_text": consolidated_answer,
            "answer_type": "Consolidated",
            "character_count": len(consolidated_answer),
            "technical_level": 1,
            "client_id": client_id,
            "rfp_id": rfp_id,
        }
        a_ins = supabase.table("client_answers").insert(a_row).execute()
        new_a_id = (a_ins.data or [{}])[0].get("id")

        # Create mapping for consolidated QA
        if new_q_id and new_a_id:
            supabase.table("client_question_answer_mappings").insert({
                "question_id": new_q_id,
                "answer_id": new_a_id,
                "confidence_score": 1.0,
                "context_requirements": None,
                "stakeholder_approved": True,
            }).execute()

        # Insert summary record
        s_row = {
            "summary_text": consolidated_answer,
            "summary_type": "Consolidated",
            "character_count": len(consolidated_answer),
            "quality_score": None,
            "approved": True,
            "client_id": client_id,
            "rfp_id": rfp_id,
        }
        s_ins = supabase.table("client_summaries").insert(s_row).execute()
        new_s_id = (s_ins.data or [{}])[0].get("id")

        # Map all original questions to the summary
        if new_s_id and question_ids:
            mappings = [{"question_id": qid, "summary_id": new_s_id} for qid in question_ids]
            supabase.table("client_question_summary_mappings").insert(mappings).execute()

        return {
            "success": True,
            "new_question_id": new_q_id,
            "new_answer_id": new_a_id,
            "new_summary_id": new_s_id,
            "mapped_questions": len(question_ids)
        }

    except Exception as e:
        print(f"Error approving summary: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to approve summary: {str(e)}")


@app.post("/org/qa/score")
def score_org_answers(payload: dict = Body(default={}), x_client_key: str | None = Header(default=None, alias="X-Client-Key"), rfp_id: str | None = Header(default=None, alias="X-RFP-ID")):
    client_id = get_client_id_from_key(x_client_key)
    limit = payload.get("limit")
    reference_text = payload.get("reference_text") or ""

    # fetch mappings and join QA
    try:
        rows = supabase.table("client_question_answer_mappings").select(
            "question_id, answer_id, client_questions(original_text, client_id, rfp_id), client_answers(answer_text, quality_score, client_id, rfp_id)"
        ).execute().data or []
    except Exception as e:
        print(f"score_org_answers fetch error: {e}")
        traceback.print_exc()
        rows = []

    # filter to this org and optionally RFP
    scoped = []
    for r in rows:
        q_data = r.get("client_questions") or {}
        a_data = r.get("client_answers") or {}
        if (q_data.get("client_id") == client_id and a_data.get("client_id") == client_id and
            (rfp_id is None or (q_data.get("rfp_id") == rfp_id and a_data.get("rfp_id") == rfp_id))):
            scoped.append(r)

    if limit:
        scoped = scoped[: int(limit)]

    updated = 0
    for r in scoped:
        question_text = (r.get("client_questions") or {}).get("original_text") or ""
        answer_id = r.get("answer_id")
        answer_text = (r.get("client_answers") or {}).get("answer_text") or ""
        if not answer_id or not question_text or not answer_text:
            continue

        prompt = _build_scoring_prompt(question_text, answer_text, reference_text)
        try:
            resp = gemini.generate_content(prompt)
            resp_text = resp.text or ""
        except Exception as e:
            print(f"score_org_answers LLM error: {e}")
            traceback.print_exc()
            resp_text = ""

        score = _extract_score(resp_text)
        if score == -1:
            continue

        try:
            supabase.table("client_answers").update({"quality_score": int(score)}).eq("id", answer_id).execute()
            updated += 1
        except Exception as e:
            print(f"score_org_answers update error: {e}")
            traceback.print_exc()

    return {"updated": updated}


# --- Q&A Management routes ---

@app.put("/questions/{question_id}")
def update_question(question_id: str, payload: dict = Body(...), x_client_key: str | None = Header(default=None, alias="X-Client-Key")):
    client_id = get_client_id_from_key(x_client_key)
    updates = {k: v for k, v in payload.items() if k in ("original_text", "category")}
    if "original_text" in updates:
        updates["normalized_text"] = updates["original_text"].lower()
    try:
        res = supabase.table("client_questions").update(updates).eq("id", question_id).eq("client_id", client_id).execute()
        return {"ok": True}
    except Exception as e:
        print(f"Error updating question {question_id}: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Failed to update question")

@app.delete("/questions/{question_id}")
def delete_question(question_id: str, x_client_key: str | None = Header(default=None, alias="X-Client-Key")):
    client_id = get_client_id_from_key(x_client_key)
    try:
        supabase.table("client_questions").delete().eq("id", question_id).eq("client_id", client_id).execute()
        return {"ok": True}
    except Exception as e:
        print(f"Error deleting question {question_id}: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Failed to delete question")

@app.put("/answers/{answer_id}")
def update_answer(answer_id: str, payload: dict = Body(...), x_client_key: str | None = Header(default=None, alias="X-Client-Key")):
    client_id = get_client_id_from_key(x_client_key)
    updates = {k: v for k, v in payload.items() if k in ("answer_text", "quality_score")}
    if "answer_text" in updates:
        updates["character_count"] = len(updates["answer_text"])
    try:
        res = supabase.table("client_answers").update(updates).eq("id", answer_id).eq("client_id", client_id).execute()
        return {"ok": True}
    except Exception as e:
        print(f"Error updating answer {answer_id}: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Failed to update answer")


@app.get("/org/summaries")
def list_org_summaries(x_client_key: str | None = Header(default=None, alias="X-Client-Key"), rfp_id: str | None = Header(default=None, alias="X-RFP-ID")):
    """List summaries and their mapped questions for an organization (optionally scoped to an RFP)"""
    client_id = get_client_id_from_key(x_client_key)
    try:
        s_query = supabase.table("client_summaries").select("id, summary_text, summary_type, character_count, quality_score, created_at, rfp_id, client_id, approved").eq("client_id", client_id).eq("approved", True)
        if rfp_id:
            s_query = s_query.eq("rfp_id", rfp_id)
        summaries = s_query.order("created_at", desc=True).execute().data or []

        if not summaries:
            return {"summaries": [], "mappings": [], "questions": []}

        s_ids = [s["id"] for s in summaries]
        m_rows = supabase.table("client_question_summary_mappings").select("question_id, summary_id").in_("summary_id", s_ids).execute().data or []

        q_ids = list({m["question_id"] for m in m_rows})
        questions = []
        if q_ids:
            q_rows = supabase.table("client_questions").select("id, original_text, category, created_at, rfp_id").in_("id", q_ids).execute().data or []
            questions = q_rows

        return {"summaries": summaries, "mappings": m_rows, "questions": questions}
    except Exception as e:
        print(f"list_org_summaries error: {e}")
        traceback.print_exc()
        return {"summaries": [], "mappings": [], "questions": []}

@app.get("/org/summaries/pending")
def list_pending_summaries(x_client_key: str | None = Header(default=None, alias="X-Client-Key"), rfp_id: str | None = Header(default=None, alias="X-RFP-ID")):
    """List pending (unapproved) summaries and their mapped questions for an organization (optionally scoped to an RFP)"""
    client_id = get_client_id_from_key(x_client_key)
    try:
        s_query = supabase.table("client_summaries").select("id, summary_text, summary_type, character_count, quality_score, created_at, rfp_id, client_id, approved").eq("client_id", client_id).eq("approved", False)
        if rfp_id:
            s_query = s_query.eq("rfp_id", rfp_id)
        summaries = s_query.order("created_at", desc=True).execute().data or []

        if not summaries:
            return {"summaries": [], "mappings": [], "questions": []}

        s_ids = [s["id"] for s in summaries]
        m_rows = supabase.table("client_question_summary_mappings").select("question_id, summary_id").in_("summary_id", s_ids).execute().data or []

        q_ids = list({m["question_id"] for m in m_rows})
        questions = []
        if q_ids:
            q_rows = supabase.table("client_questions").select("id, original_text, category, created_at, rfp_id").in_("id", q_ids).execute().data or []
            questions = q_rows

        return {"summaries": summaries, "mappings": m_rows, "questions": questions}
    except Exception as e:
        print(f"list_pending_summaries error: {e}")
        traceback.print_exc()
        return {"summaries": [], "mappings": [], "questions": []}

@app.post("/org/summaries/{summary_id}/set-approval")
def set_summary_approval(summary_id: str, payload: dict = Body(...), x_client_key: str | None = Header(default=None, alias="X-Client-Key")):
    """Set approval status for a summary (approved true/false)."""
    client_id = get_client_id_from_key(x_client_key)
    approved = bool(payload.get("approved", False))
    try:
        supabase.table("client_summaries").update({"approved": approved}).eq("id", summary_id).eq("client_id", client_id).execute()
        return {"ok": True, "approved": approved}
    except Exception as e:
        print(f"set_summary_approval error: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Failed to update summary approval")
@app.delete("/answers/{answer_id}")
def delete_answer(answer_id: str, x_client_key: str | None = Header(default=None, alias="X-Client-Key")):
    client_id = get_client_id_from_key(x_client_key)
    try:
        supabase.table("client_answers").delete().eq("id", answer_id).eq("client_id", client_id).execute()
        return {"ok": True}
    except Exception as e:
        print(f"delete_answer error: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Failed to delete answer")


# --- Job Management System ---

def estimate_processing_time(file_size: int, job_type: str) -> int:
    """Estimate processing time in minutes based on file size and job type"""
    if job_type == "process_rfp":
        # Rough estimate: 1 minute per 100KB for RFP processing
        return max(5, min(30, file_size // 102400))
    elif job_type == "extract_qa":
        # Rough estimate: 1 minute per 200KB for QA extraction
        return max(3, min(20, file_size // 204800))
    return 10


def _estimate_minutes_from_chars(file_bytes: bytes, job_type: str) -> int:
    """Estimate processing time based on total character count across all sheets.
    Falls back to file-size estimation on error.
    """
    try:
        bio = io.BytesIO(file_bytes)
        xls = pd.ExcelFile(bio)
        total_chars = 0
        for sheet in xls.sheet_names:
            try:
                df = pd.read_excel(xls, sheet_name=sheet, header=None, engine="openpyxl")
                if df is None or df.empty:
                    continue
                # Drop fully empty rows/cols, then convert to strings
                df = df.dropna(how='all')
                df = df.dropna(axis=1, how='all')
                if df.empty:
                    continue
                # Convert to strings and sum lengths; use tab as cell sep and newline as row sep
                # to approximate real characters analyzed
                texts = df.astype(str).apply(lambda row: "\t".join(row.values), axis=1)
                sheet_chars = sum(len(s) for s in texts)
                total_chars += sheet_chars
            except Exception as e:
                print(f"Error in _estimate_minutes_from_chars for sheet {sheet}: {e}")
                traceback.print_exc()
                continue

        # Base rates per job type (characters per minute)
        if job_type == "process_rfp":
            chars_per_min = 12000  # more conservative to reflect real processing
            min_minutes, max_minutes = 5, 90
        else:  # extract_qa
            chars_per_min = 20000
            min_minutes, max_minutes = 3, 60

        estimated = int(max(min_minutes, min(max_minutes, (total_chars // max(1, chars_per_min)) + 1)))
        # Ensure at least 1 minute if tiny
        return max(1, estimated)
    except Exception as e:
        print(f"Error in _estimate_minutes_from_chars: {e}")
        traceback.print_exc()
        # Fallback: use previous file-size based estimator
        return estimate_processing_time(len(file_bytes), job_type)

def process_excel_file_obj(file_obj: io.BytesIO, filename: str, client_id: str, rfp_id: str = None, job_id: str = None) -> tuple[io.BytesIO, int, int]:
    """
    Process Excel file, add AI answers, and return the processed file as BytesIO,
    along with counts of processed sheets and questions.
    """
    print(f"DEBUG: process_excel_file_obj started for {filename} (Job ID: {job_id})")
    start_time = time.time()
    max_processing_time = 1800  # 30 minutes max processing time
    
    _processed_sheets_count = 0
    _total_questions_processed = 0

    try:
        original_bytes = file_obj.getvalue()

        wb = openpyxl.load_workbook(io.BytesIO(original_bytes))
        xls = pd.ExcelFile(io.BytesIO(original_bytes))

        total_sheets = len(wb.sheetnames)
        
        ai_col = None
        review_col = None
        
        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            try:
                if time.time() - start_time > max_processing_time:
                    raise Exception(f"Processing timeout: exceeded {max_processing_time/60:.1f} minutes")
                
                update_job_progress(job_id, 15 + int(sheet_idx * 70 / total_sheets / 2),
                                    f"Processing sheet {sheet_idx + 1}/{total_sheets}: {sheet_name}")
                print(f"DEBUG: Processing sheet {sheet_idx + 1}/{total_sheets}: {sheet_name}")
                ws = wb[sheet_name]

                df = pd.read_excel(xls, sheet_name=sheet_name, header=None, engine="openpyxl")
                
                if df is not None:
                    try:
                        df = df.dropna(how='all')
                        df = df.dropna(axis=1, how='all')
                    except Exception as e:
                        print(f"WARN: Error dropping empty rows/cols in sheet {sheet_name}: {e}")
                        traceback.print_exc()

                if df is None or df.empty:
                    print(f"DEBUG: Sheet {sheet_name} is empty after cleanup, skipping.")
                    _processed_sheets_count += 1
                    continue
                    
                # Initialize AI Answer and Review Status columns
                ai_col = find_first_empty_data_column(ws)
                ws.cell(row=1, column=ai_col, value="AI Answers").alignment = Alignment(wrap_text=True)
                review_col = ai_col + 1
                ws.cell(row=1, column=review_col, value="Review Status").alignment = Alignment(wrap_text=True)

                # Set default column widths for new columns
                ws.column_dimensions[get_column_letter(ai_col)].width = 20
                ws.column_dimensions[get_column_letter(review_col)].width = 15

                # ============================================================
                # PROCESSING FLOW - Direct Row Mapping (No Searching!)
                # ============================================================
                # 1. Extract rows as arrays (like XLSX.utils.sheet_to_json)
                #    Row 5 → ["What", "is", "your", "approach", "?"]
                #
                # 2. Create RowWithNumber objects: {rowNumber, rowData}
                #    {"rowNumber": 5, "rowData": ["What", "is", ...]}
                #
                # 3. Send to AI in batches of 50
                #    detectQuestionsInBatch([{"rowNumber": 5, ...}, ...])
                #
                # 4. AI returns with SAME row numbers:
                #    [{"rowNumber": 5, "question": "What is your approach?"}]
                #
                # 5. Generate answers (still using row number):
                #    {5: {"answer": "Our approach...", "review_status": "..."}}
                #
                # 6. Place answer in EXACT row (NO SEARCHING!):
                #    ws.cell(row=5, column=ai_col, value="Our approach...")
                #
                # Result: Question in Row 5 → Answer in Row 5 ✓
                # ============================================================
                
                max_row = ws.max_row or 1
                min_data_row = 2  # Start from row 2 (skip header)
                batch_size = 50  # Process 50 rows per batch with structured output
                total_questions_in_sheet = 0
                
                # Step 1 & 2: Collect all rows with their data (TypeScript: parseXlsx equivalent)
                # IMPORTANT: Start from min_data_row (row 2) to SKIP HEADER ROW (row 1)
                rows_with_numbers = []
                
                print(f"DEBUG: Scanning rows {min_data_row} to {max_row} in sheet {sheet_name}")
                
                for row_num in range(min_data_row, max_row + 1):
                    # SAFETY CHECK: Never process row 1 (header)
                    if row_num < min_data_row:
                        print(f"CRITICAL: Skipping row {row_num} - below minimum data row!")
                        continue
                    
                    # Check if row has any content
                    row_has_content = False
                    for col in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row=row_num, column=col).value
                        if cell_value is not None and str(cell_value).strip():
                            row_has_content = True
                            break
                    
                    if row_has_content:
                        row_data = _extract_row_data(ws, row_num)
                        rows_with_numbers.append({
                            "rowNumber": row_num,
                            "rowData": row_data
                        })
                
                total_rows_to_process = len(rows_with_numbers)
                print(f"DEBUG: Collected {total_rows_to_process} non-empty rows from sheet {sheet_name}")
                print(f"DEBUG: Row range being processed: {min_data_row} to {max_row}")
                
                # Validate that no row 1 was included
                invalid_rows = [r["rowNumber"] for r in rows_with_numbers if r["rowNumber"] < min_data_row]
                if invalid_rows:
                    print(f"CRITICAL ERROR: Header rows detected in processing batch: {invalid_rows}. Removing them!")
                    rows_with_numbers = [r for r in rows_with_numbers if r["rowNumber"] >= min_data_row]
                
                # Step 3: Process rows in batches to detect questions
                # (TypeScript: await detectQuestionsInBatch(rows))
                all_detected_questions = []
                for batch_start in range(0, total_rows_to_process, batch_size):
                    batch_end = min(batch_start + batch_size, total_rows_to_process)
                    batch_rows = rows_with_numbers[batch_start:batch_end]
                    
                    # Update progress
                    progress_pct = 15 + int(sheet_idx * 70 / total_sheets / 2) + int(batch_start * 35 / total_rows_to_process / total_sheets)
                    update_job_progress(job_id, progress_pct,
                                        f"Detecting questions in rows {batch_start + 1}-{batch_end} of {total_rows_to_process} in sheet {sheet_name}")
                    
                    print(f"DEBUG: Detecting questions in batch {batch_start // batch_size + 1}: rows {batch_start + 1}-{batch_end}")
                    
                    # Call detect_questions_in_batch (TypeScript: detectQuestionsInBatch)
                    detected = detect_questions_in_batch(batch_rows)
                    
                    # VALIDATE: Check AI returned row numbers are within valid range
                    if detected:
                        detected_row_nums = [q.get("rowNumber", 0) for q in detected]
                        print(f"DEBUG: AI returned row numbers: {detected_row_nums}")
                        
                        # Check for invalid row numbers
                        invalid_detections = [q for q in detected if q.get("rowNumber", 0) < min_data_row or q.get("rowNumber", 0) > max_row]
                        if invalid_detections:
                            print(f"WARNING: AI returned {len(invalid_detections)} questions with invalid row numbers. Filtering them out.")
                            for inv in invalid_detections:
                                print(f"  - Invalid row {inv.get('rowNumber')}: '{inv.get('question', '')[:50]}...'")
                            # Remove invalid detections
                            detected = [q for q in detected if min_data_row <= q.get("rowNumber", 0) <= max_row]
                    
                    all_detected_questions.extend(detected)
                    print(f"DEBUG: Found {len(detected)} VALID questions in this batch")
                
                print(f"DEBUG: Total {len(all_detected_questions)} questions detected in sheet {sheet_name}")
                
                # Final validation of all detected questions
                if all_detected_questions:
                    all_row_nums = [q.get("rowNumber", 0) for q in all_detected_questions]
                    print(f"DEBUG: All detected row numbers: {sorted(set(all_row_nums))}")
                    
                    # Verify no header rows
                    header_violations = [q for q in all_detected_questions if q.get("rowNumber", 0) < min_data_row]
                    if header_violations:
                        print(f"CRITICAL ERROR: Found {len(header_violations)} questions in header rows!")
                        for hv in header_violations:
                            print(f"  - Row {hv.get('rowNumber')}: '{hv.get('question', '')[:50]}...'")
                        # Remove header violations
                        all_detected_questions = [q for q in all_detected_questions if q.get("rowNumber", 0) >= min_data_row]
                        print(f"DEBUG: After removing header violations: {len(all_detected_questions)} questions remain")
                
                # Step 4 & 5: Generate answers for detected questions
                if all_detected_questions:
                    update_job_progress(job_id, 15 + int((sheet_idx + 0.5) * 70 / total_sheets),
                                        f"Generating answers for {len(all_detected_questions)} questions in sheet {sheet_name}")
                    
                    # Pass validation parameters to ensure row numbers are correct
                    answers_dict = process_detected_questions_batch(
                        all_detected_questions, 
                        client_id, 
                        rfp_id,
                        min_valid_row=min_data_row,  # Row 2 minimum (never row 1)
                        max_valid_row=max_row          # Maximum row in sheet
                    )
                    
                    print(f"DEBUG: Generated {len(answers_dict)} answers for sheet {sheet_name}")
                    
                    # Step 6: Write answers to EXACT rowNumber
                    # ============================================================
                    # TypeScript equivalent from xlsxService.ts:
                    #   sheet.data[result.rowNumber].push(answer, reviewStatus)
                    #   OR: sheet.data[result.rowNumber][aiColumnIndex] = answer
                    #
                    # Python equivalent:
                    #   ws.cell(row=result.rowNumber, column=ai_col, value=answer)
                    #
                    # Key: Direct row number mapping - no fuzzy search, no resolution
                    # ============================================================
                    
                    # ===========================================================
                    # DIRECT ROW MAPPING - NO SEARCHING NEEDED
                    # ===========================================================
                    # The AI already told us which row contains each question:
                    #   Input:  {"rowNumber": 5, "rowData": ["What", "is", ...]}
                    #   Output: {"rowNumber": 5, "question": "What is..."}
                    #   Action: Place answer directly in row 5
                    #
                    # This is MUCH MORE RELIABLE than the old approach which:
                    # - Sent entire sheet text
                    # - Got questions with uncertain row numbers
                    # - Had to search/match to find the right row
                    # ===========================================================
                    
                    # Validate and place each answer IN THE EXACT ROW from AI
                    for row_num, answer_data in answers_dict.items():
                        # TRIPLE VALIDATION: Ensure safe answer placement
                        if row_num < min_data_row:
                            print(f"CRITICAL ERROR: Attempted to place answer in row {row_num} (below min {min_data_row}). SKIPPING to prevent header corruption!")
                            continue
                        
                        if row_num > max_row:
                            print(f"ERROR: Attempted to place answer in row {row_num} (above max {max_row}). SKIPPING!")
                            continue
                        
                        if row_num == 1:
                            print(f"CRITICAL ERROR: Attempted to place answer in HEADER ROW (row 1). SKIPPING!")
                            continue
                        
                        # Safe to place answer - DIRECT ROW MAPPING
                        try:
                            answer_text = answer_data.get("answer", "")
                            review_status = answer_data.get("review_status", "")
                            question_text = answer_data.get("question", "")
                            
                            # Place answer in THE SAME ROW where AI found the question
                            ws.cell(row=row_num, column=ai_col, value=answer_text)
                            ws.cell(row=row_num, column=review_col, value=review_status)
                            total_questions_in_sheet += 1
                            
                            print(f"✓ Row {row_num}: Placed answer for '{question_text[:60]}...'")
                        except Exception as place_error:
                            print(f"ERROR: Failed to place answer in row {row_num}: {place_error}")
                            traceback.print_exc()
                            continue
                
                print(f"\n{'='*80}")
                print(f"SHEET SUMMARY: {sheet_name}")
                print(f"  - Total data rows scanned: {total_rows_to_process}")
                print(f"  - Questions detected by AI: {len(all_detected_questions)}")
                print(f"  - Answers successfully placed: {total_questions_in_sheet}")
                print(f"  - Row range: {min_data_row} to {max_row}")
                if total_questions_in_sheet > 0:
                    placed_rows = [r for r, _ in answers_dict.items() if min_data_row <= r <= max_row]
                    print(f"  - Rows with answers: {sorted(placed_rows)}")
                print(f"{'='*80}\n")
                
                _total_questions_processed += total_questions_in_sheet
                
                del df
                gc.collect()
                _processed_sheets_count += 1
                
            except Exception as sheet_error:
                print(f"ERROR: Failed to process sheet {sheet_name}: {sheet_error}")
                traceback.print_exc()
                continue

        update_job_progress(job_id, 90, "Saving processed Excel file...")
        
        # ============================================================
        # TypeScript equivalent from xlsxService.ts:
        #   const worksheet = XLSX.utils.aoa_to_sheet(sheet.data);
        #   XLSX.utils.book_append_sheet(workbook, worksheet, sheet.name);
        #   XLSX.writeFile(workbook, newFilename);
        #
        # Python equivalent (openpyxl):
        #   wb.save(output) - saves the modified workbook
        #
        # The workbook already has answers in the correct rows,
        # now we just format and save.
        # ============================================================
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Adjust column widths for new AI Answer and Review Status columns
            if ai_col and ai_col <= ws.max_column:
                ws.column_dimensions[get_column_letter(ai_col)].width = 60
            if review_col and review_col <= ws.max_column:
                ws.column_dimensions[get_column_letter(review_col)].width = 25
            
            # Enable text wrapping for the AI answer column and review status
            for col_to_wrap in [ai_col, review_col]:
                if col_to_wrap and col_to_wrap <= ws.max_column:
                    for row in range(2, ws.max_row + 1): # Start from row 2 for data cells
                        cell = ws.cell(row=row, column=col_to_wrap)
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Save workbook (TypeScript: XLSX.writeFile)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        print(f"DEBUG: process_excel_file_obj completed for {filename} in {time.time() - start_time:.1f}s.")
        return output, _processed_sheets_count, _total_questions_processed
        
    except Exception as e:
        processing_time = time.time() - start_time
        error_msg = f"Excel processing failed after {processing_time:.1f}s: {str(e)}"
        print(f"ERROR: process_excel_file_obj error for {filename}: {error_msg}")
        traceback.print_exc()
        raise Exception(error_msg)

def update_job_progress(job_id: str, progress: int, current_step: str, result_data: dict = None):
    """Update job progress in database with retry logic"""
    max_retries = 3
    for attempt in range(max_retries):
        try:
            updates = {
                "progress_percent": progress,
                "current_step": current_step,
                "last_updated": datetime.now().isoformat()
            }
            if result_data:
                updates["result_data"] = result_data
            if progress == 100:
                updates["status"] = "completed"
                updates["completed_at"] = datetime.now().isoformat()
            elif progress == -1: # Use -1 for explicit failure
                updates["status"] = "failed"
                updates["completed_at"] = datetime.now().isoformat()
            
            print(f"DEBUG: Updating job {job_id} - Progress: {progress}%, Step: {current_step}")
            # Ensure the client is active for Supabase calls, re-create if necessary
            global supabase
            if supabase is None:
                supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
            
            supabase.table("client_jobs").update(updates).eq("id", job_id).execute()
            return  # Success, exit retry loop
        except Exception as e:
            print(f"ERROR: Error updating job progress {job_id} (attempt {attempt + 1}/{max_retries}): {e}")
            traceback.print_exc()
            if attempt < max_retries - 1:
                time.sleep(1)  # Wait 1 second before retry
            else:
                print(f"CRITICAL ERROR: Failed to update job progress for {job_id} after {max_retries} attempts.")

def process_rfp_background(job_id: str, file_content: bytes, file_name: str, client_id: str, rfp_id: str):
    """Background RFP processing function using the working code"""
    print(f"DEBUG: Background RFP processing started for job {job_id}")
    start_time = time.time()
    
    try:
        update_job_progress(job_id, 10, "Starting RFP processing: Loading file...")
        
        # Check file size to prevent memory issues
        file_size_mb = len(file_content) / (1024 * 1024)
        if file_size_mb > 50:  # 50MB limit
            raise Exception(f"File too large: {file_size_mb:.1f}MB. Maximum allowed: 50MB")
        
        print(f"DEBUG: Processing file {file_name} ({file_size_mb:.1f}MB)")
        
        file_obj = io.BytesIO(file_content)
        processed_output, processed_sheets_count, total_questions_processed = process_excel_file_obj(file_obj, file_name, client_id, rfp_id, job_id=job_id)
        
        processed_content = processed_output.getvalue()
        
        update_job_progress(job_id, 95, "Finalizing and storing processed file...")
        
        # Store both original and processed files for comparison
        import base64
        processed_file_b64 = base64.b64encode(processed_content).decode('utf-8')
        original_file_b64 = base64.b64encode(file_content).decode('utf-8')
        
        result_data = {
            "file_name": f"processed_{file_name}",
            "file_size": len(processed_content),
            "processing_completed": True,
            "processing_time_seconds": int(time.time() - start_time),
            "processed_file": processed_file_b64,  # Store as base64
            "original_file": original_file_b64,    # Store original for comparison
            "sheets_processed": processed_sheets_count,
            "total_questions_processed": total_questions_processed
        }
        
        update_job_progress(job_id, 100, "RFP processing completed successfully!", result_data)
        print(f"DEBUG: Background RFP processing completed successfully for job {job_id} in {time.time() - start_time:.1f}s")
                
    except Exception as e:
        processing_time = time.time() - start_time
        error_msg = f"Processing failed after {processing_time:.1f}s: {str(e)}"
        print(f"ERROR: RFP processing background error for job {job_id}: {error_msg}")
        traceback.print_exc()
        update_job_progress(job_id, -1, error_msg)

def extract_qa_background(job_id: str, file_content: bytes, file_name: str, client_id: str, rfp_id: str):
    """Background QA extraction function"""
    print(f"DEBUG: Background QA extraction started for job {job_id}")
    try:
        update_job_progress(job_id, 10, "Starting QA extraction: Loading file...")
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file.write(file_content)
            tmp_path = tmp_file.name
        
        try:
            update_job_progress(job_id, 20, "Loading Excel file for QA extraction...")
            xls = pd.ExcelFile(tmp_path)
            
            total_sheets = len(xls.sheet_names)
            processed_sheets = 0
            extracted_pairs = []
            
            for sheet_idx, sheet in enumerate(xls.sheet_names):
                progress_start_sheet = 20 + (sheet_idx * 70 // total_sheets)
                progress_end_sheet = 20 + ((sheet_idx + 1) * 70 // total_sheets)
                
                update_job_progress(job_id, progress_start_sheet, f"Extracting from sheet {sheet_idx + 1}/{total_sheets}: {sheet}")
                
                
                df = pd.read_excel(tmp_path, sheet_name=sheet, header=None)
                if df.empty:
                    print(f"DEBUG: Sheet {sheet} is empty, skipping for QA extraction.")
                    processed_sheets += 1
                    continue
                
                sheet_csv = df.to_csv(index=False, header=False)
                pairs = _extract_qa_pairs(sheet_csv)
                # Retry up to 2 additional times if no pairs extracted
                retry_attempt = 0
                while (not pairs or len(pairs) == 0) and retry_attempt < 2:
                    retry_attempt += 1
                    print(f"DEBUG: No Q&A pairs extracted from sheet {sheet}. Retrying attempt {retry_attempt}/2...")
                pairs = _extract_qa_pairs(sheet_csv)
                
                for p_idx, p in enumerate(pairs):
                    extracted_pairs.append({
                        "question": p.get("question", ""),
                        "answer": p.get("answer", ""),
                        "category": p.get("category", "Other"),
                        "sheet": sheet
                    })
                
                processed_sheets += 1
            
            update_job_progress(job_id, 90, f"Saving {len(extracted_pairs)} extracted Q&A pairs to database...")
            print(f"DEBUG: Saving {len(extracted_pairs)} extracted Q&A pairs to database for job {job_id}.")
            
            created_count = 0
            for p_idx, p in enumerate(extracted_pairs):
                q = p.get("question", "").strip()
                a = p.get("answer", "").strip()
                c = p.get("category", "Other").strip() or "Other"
                if q and a and _insert_qa_pair(client_id, q, a, c, rfp_id):
                    created_count += 1
            
            def _build_ai_groups_for_job(client_id_param: str, rfp_id_param: str | None):
                try:
                    q_query = supabase.table("client_questions").select("id, original_text, rfp_id").eq("client_id", client_id_param)
                    if rfp_id_param:
                        q_query = q_query.eq("rfp_id", rfp_id_param)
                    questions_local = q_query.order("created_at", desc=True).execute().data or []

                    if not questions_local:
                        return {"groups": [], "message": "No questions found"}

                    q_ids_local = [q["id"] for q in questions_local]
                    m_rows_local = supabase.table("client_question_answer_mappings").select("question_id, answer_id").in_("question_id", q_ids_local).execute().data or []
                    q_to_a_local = {m["question_id"]: m["answer_id"] for m in m_rows_local}

                    a_ids_local = list({aid for aid in q_to_a_local.values() if aid})
                    a_rows_local = []
                    if a_ids_local:
                        a_rows_local = supabase.table("client_answers").select("id, answer_text").in_("id", a_ids_local).execute().data or []
                    a_map_local = {a["id"]: a.get("answer_text", "") for a in a_rows_local}

                    qa_lines_local = []
                    for q in questions_local:
                        qid = q["id"]
                        qtext = (q.get("original_text") or "").strip()
                        atext = (a_map_local.get(q_to_a_local.get(qid)) or "").strip()
                        if not qtext:
                            continue
                        qa_lines_local.append({"id": qid, "q": qtext, "a": atext})

                    if not qa_lines_local:
                        return {"groups": [], "message": "No Q&A pairs available"}

                    qa_text_local = "\n".join([f"ID:{row['id']}\nQ:{row['q']}\nA:{row['a']}" for row in qa_lines_local])
                    prompt_local = f"""
You will receive a list of Q&A pairs for a single client (and optionally a specific RFP). Group semantically similar questions together and produce a consolidated Q&A for each group.

Return ONLY strict JSON with this structure:
{{
  "groups": [
    {{
      "question_ids": ["<id>", "<id>", ...],
      "consolidated_question": "string",
      "consolidated_answer": "string"
    }}
  ]
}}

Rules:
1) "question_ids" must be the exact IDs provided.
2) Every ID used must exist in the input. Do not invent IDs.
3) Do not include any text before or after the JSON.

Q&A LIST:
{qa_text_local}
"""
                    try:
                        response_local = gemini.generate_content(
                            prompt_local,
                            safety_settings={
                                HarmCategory.HARM_CATEGORY_HARASSMENT: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                                HarmCategory.HARM_CATEGORY_HATE_SPEECH: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                                HarmCategory.HARM_CATEGORY_SEXUALLY_EXPLICIT: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                                HarmCategory.HARMS_CATEGORY_DANGEROUS_CONTENT: HarmBlockThreshold.BLOCK_ONLY_HIGH,
                            },
                        )
                        text_local = (response_local.text or "").strip()
                        s_local, e_local = text_local.find("{"), text_local.rfind("}")
                        if s_local == -1 or e_local == -1:
                            return {"groups": [], "message": "LLM returned no JSON"}
                        data_local = json.loads(text_local[s_local:e_local+1])

                        raw_groups_local = data_local.get("groups") or []
                        valid_ids_local = {str(x["id"]) for x in questions_local}
                        groups_local = []
                        for g in raw_groups_local:
                            qids_local = [str(x) for x in (g.get("question_ids") or []) if str(x) in valid_ids_local]
                            if not qids_local:
                                continue
                            groups_local.append({
                                "question_ids": qids_local,
                                "consolidated_question": (g.get("consolidated_question") or "").strip(),
                                "consolidated_answer": (g.get("consolidated_answer") or "").strip(),
                            })
                        return {"groups": groups_local}
                    except Exception as e:
                        print(f"_build_ai_groups_for_job LLM error: {e}")
                        traceback.print_exc()
                        return {"groups": [], "message": "AI grouping failed"}
                except Exception as e:
                    print(f"_build_ai_groups_for_job error: {e}")
                    traceback.print_exc()
                    return {"groups": [], "message": "AI grouping error"}

            ai_groups_result = _build_ai_groups_for_job(client_id, rfp_id)

            try:
                for grp in ai_groups_result.get("groups", []):
                    cq = (grp.get("consolidated_question") or "").strip()
                    ca = (grp.get("consolidated_answer") or "").strip()
                    qids = grp.get("question_ids") or []
                    if not cq or not ca or not qids:
                        continue
                    s_ins = supabase.table("client_summaries").insert({
                        "summary_text": ca,
                        "summary_type": "Consolidated",
                        "character_count": len(ca),
                        "quality_score": None,
                        "approved": False,
                        "client_id": client_id,
                        "rfp_id": rfp_id,
                    }).execute()
                    s_id = (s_ins.data or [{}])[0].get("id")
                    if s_id:
                        mappings = [{"question_id": qid, "summary_id": s_id} for qid in qids]
                        supabase.table("client_question_summary_mappings").insert(mappings).execute()
            except Exception as e:
                print(f"WARN: Failed to persist pending summaries: {e}")
                traceback.print_exc()

            result_data = {
                "extracted_pairs_count": len(extracted_pairs), 
                "created_pairs_count": created_count,          
                "total_sheets_processed": processed_sheets,
                "ai_groups": ai_groups_result.get("groups", []),
                "ai_groups_count": len(ai_groups_result.get("groups", [])),
            }
            
            update_job_progress(job_id, 100, "QA extraction completed successfully!", result_data)
            print(f"DEBUG: Background QA extraction completed successfully for job {job_id}")
            
        finally:
            if os.path.exists(tmp_path):
                try:
                    os.unlink(tmp_path)
                    print(f"DEBUG: Cleaned up temporary file: {tmp_path}")
                except PermissionError:
                    print(f"DEBUG: Could not delete temp file {tmp_path} - file may be in use")
                except Exception as e:
                    print(f"DEBUG: Error cleaning up temp file: {e}")
                    traceback.print_exc()
                
    except Exception as e:
        print(f"ERROR: QA extraction background error for job {job_id}: {e}")
        traceback.print_exc()
        update_job_progress(job_id, -1, f"Extraction failed: {str(e)}")

# Google Drive Integration Functions
def get_drive_service(access_token: str):
    """Create Google Drive service with access token"""
    credentials = Credentials(token=access_token)
    return build('drive', 'v3', credentials=credentials)

def find_or_create_folder(service, folder_name: str, parent_folder_id: str = None) -> str:
    """Find existing folder or create new one"""
    try:
        query = f"name='{folder_name}' and mimeType='application/vnd.google-apps.folder'"
        if parent_folder_id:
            query += f" and '{parent_folder_id}' in parents"
        else:
            query += " and 'root' in parents"
        
        results = service.files().list(q=query, fields="files(id, name)").execute()
        files = results.get('files', [])
        
        if files:
            return files[0]['id']
        
        folder_metadata = {
            'name': folder_name,
            'mimeType': 'application/vnd.google-apps.folder'
        }
        if parent_folder_id:
            folder_metadata['parents'] = [parent_folder_id]
        
        folder = service.files().create(body=folder_metadata, fields='id').execute()
        return folder.get('id')
        
    except Exception as e:
        print(f"Error finding/creating folder: {e}")
        traceback.print_exc()
        return None

def upload_file_to_drive(service, file_content: bytes, filename: str, folder_id: str, mime_type: str = 'application/octet-stream') -> str:
    """Upload file to Google Drive folder"""
    try:
        file_metadata = {
            'name': filename,
            'parents': [folder_id]
        }
        
        media = MediaIoBaseUpload(io.BytesIO(file_content), mimetype=mime_type, resumable=True)
        file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        return file.get('id')
        
    except Exception as e:
        print(f"Error uploading file to Drive: {e}")
        traceback.print_exc()
        return None

def setup_drive_folders(access_token: str, client_name: str) -> dict:
    """Setup folder structure in Google Drive"""
    try:
        service = get_drive_service(access_token)
        
        client_folder_id = find_or_create_folder(service, f"Your_RFP_{client_name}")
        if not client_folder_id:
            return None
        
        processed_folder_id = find_or_create_folder(service, "Processed Files", client_folder_id)
        unprocessed_folder_id = find_or_create_folder(service, "Unprocessed Files", client_folder_id)
        
        return {
            "client_folder_id": client_folder_id,
            "processed_folder_id": processed_folder_id,
            "unprocessed_folder_id": unprocessed_folder_id
        }
        
    except Exception as e:
        print(f"Error setting up Drive folders: {e}")
        traceback.print_exc()
        return None

def create_rfp_from_filename(client_id: str, filename: str) -> str:
    """Auto-create RFP record from filename"""
    rfp_name = filename.rsplit('.', 1)[0]
    rfp_name = rfp_name.replace('_', ' ').replace('-', ' ')
    rfp_name = ' '.join(word.capitalize() for word in rfp_name.split())
    
    try:
        existing_rfps_res = supabase.table("client_rfps").select("id, name").eq("client_id", client_id).execute()
        existing_rfps = existing_rfps.data or []
    except Exception as e:
        print(f"Error checking for existing RFPs: {e}")
        traceback.print_exc()
        existing_rfps = []
    
    rfp_id = None
    for rfp in existing_rfps:
        if rfp["name"].lower() == rfp_name.lower():
            rfp_id = rfp["id"]
            print(f"DEBUG: Found existing RFP '{rfp_name}' with ID: {rfp_id}")
            break
    
    if not rfp_id:
        rfp_data = {
            "client_id": client_id,
            "name": rfp_name,
            "description": f"Auto-created from uploaded file: {filename}",
            "created_at": datetime.now().isoformat()
        }
        try:
            rfp_result = supabase.table("client_rfps").insert(rfp_data).execute()
            rfp_id = rfp_result.data[0]["id"]
            print(f"DEBUG: Created new RFP '{rfp_name}' with ID: {rfp_id}")
        except Exception as e:
            print(f"Error creating new RFP: {e}")
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Failed to auto-create RFP from filename: {str(e)}")
    
    return rfp_id

@app.post("/jobs/submit")
async def submit_job(
    file: UploadFile = File(...),
    job_type: str = Form(...),
    x_client_key: str | None = Header(default=None, alias="X-Client-Key")
):
    """Submit a job for background processing - uses worker process for long-running tasks"""
    try:
        print(f"=== /jobs/submit ENDPOINT CALLED ===")
        print(f"Received job submission: file={file.filename}, job_type={job_type}")
        
        if not file.filename:
            print("ERROR: No file filename provided")
            raise HTTPException(status_code=400, detail="No file provided")
        
        if not job_type:
            print("ERROR: No job type provided")
            raise HTTPException(status_code=400, detail="No job type provided")
        
        client_id = get_client_id_from_key(x_client_key)
        print(f"Resolved client_id: {client_id}")
        
        if job_type not in ["process_rfp", "extract_qa"]:
            print(f"ERROR: Invalid job type: {job_type}")
            raise HTTPException(status_code=400, detail=f"Invalid job type: {job_type}. Must be 'process_rfp' or 'extract_qa'")
        
        print("Validation passed, proceeding with job creation...")
    except HTTPException as he:
        print(f"HTTPException in submit_job: {he.detail}")
        traceback.print_exc()
        raise
    except Exception as e:
        print(f"ERROR: Error in submit_job before background task: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
    
    rfp_id = create_rfp_from_filename(client_id, file.filename)
    print(f"Created RFP with ID: {rfp_id}")
    
    print("Reading file content...")
    file_content = await file.read()
    file_size = len(file_content)
    print(f"File size: {file_size} bytes")
    
    file_size_mb = file_size / (1024 * 1024)
    if file_size_mb > 50:
        raise HTTPException(status_code=400, detail=f"File too large: {file_size_mb:.1f}MB. Maximum allowed: 50MB")
    
    estimated_minutes = _estimate_minutes_from_chars(file_content, job_type)
    estimated_completion = datetime.now() + timedelta(minutes=estimated_minutes)
    
    import base64
    file_content_b64 = base64.b64encode(file_content).decode('utf-8')
    
    print("Creating job record...")
    job_data = {
        "client_id": client_id,
        "rfp_id": rfp_id,
        "job_type": job_type,
        "status": "pending",
        "file_name": file.filename,
        "file_size": file_size,
        "progress_percent": 0,
        "current_step": "Job queued for processing",
        "estimated_completion": estimated_completion.isoformat(),
        "created_at": datetime.now().isoformat(),
        "job_data": {
            "file_content": file_content_b64,
            "file_name": file.filename,
            "client_id": client_id,
            "rfp_id": rfp_id
        }
    }
    print(f"Job data created, file content encoded")
    
    try:
        job_result = supabase.table("client_jobs").insert(job_data).execute()
        job_id = job_result.data[0]["id"]
        print(f"Created job with ID: {job_id}")
    except Exception as e:
        print(f"Error inserting job data: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to create job record: {str(e)}")
    
    result = {
        "job_id": job_id, 
        "rfp_id": rfp_id, 
        "estimated_minutes": estimated_minutes, 
        "status": "submitted",
        "message": "Job submitted successfully. Processing will begin shortly."
    }
    print(f"Returning job submission result for job_id={result.get('job_id')}")
    return result

@app.get("/jobs")
def list_jobs(x_client_key: str | None = Header(default=None, alias="X-Client-Key")):
    global supabase
    """List all jobs for a client with retry logic"""
    print(f"=== /jobs ENDPOINT CALLED ===")
    print("Jobs endpoint called")
    client_id = get_client_id_from_key(x_client_key)
    max_retries = 3
    for attempt in range(max_retries):
        try:
            res = supabase.table("client_jobs").select("*").eq("client_id", client_id).order("created_at", desc=True).execute()
            jobs = res.data or []
            for job in jobs[:2]:
                print(f"DEBUG: Job {job.get('id', 'unknown')} - status: {job.get('status')} progress: {job.get('progress_percent')}")
            return {"jobs": jobs}
        except Exception as e:
            print(f"ERROR: Error fetching jobs (attempt {attempt + 1}/{max_retries}): {e}")
            traceback.print_exc()
            try:
                supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
            except Exception as reinit_err:
                print(f"Supabase re-init failed in /jobs: {reinit_err}")
                traceback.print_exc()
            if attempt < max_retries - 1:
                time.sleep(1)
            else:
                print(f"Failed to fetch jobs after {max_retries} attempts")
                return {"jobs": [], "error": "Database connection failed"}

# Google Drive API endpoints
@app.post("/drive/setup")
def setup_drive_folders_endpoint(
    payload: dict = Body(...),
    x_client_key: str | None = Header(default=None, alias="X-Client-Key")
):
    """Setup Google Drive folders for client"""
    client_id = get_client_id_from_key(x_client_key)
    access_token = payload.get("access_token")
    client_name = payload.get("client_name", "Client")
    
    if not access_token:
        raise HTTPException(status_code=400, detail="Access token required")
    
    try:
        folder_structure = setup_drive_folders(access_token, client_name)
        if folder_structure:
            return {"success": True, "folders": folder_structure}
        else:
            raise HTTPException(status_code=500, detail="Failed to setup Drive folders")
    except Exception as e:
        print(f"Drive setup error: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Drive setup failed: {str(e)}")

@app.post("/drive/upload")
def upload_to_drive_endpoint(
    payload: dict = Body(...),
    x_client_key: str | None = Header(default=None, alias="X-Client-Key")
):
    """Upload file to Google Drive"""
    client_id = get_client_id_from_key(x_client_key)
    access_token = payload.get("access_token")
    file_content = payload.get("file_content")
    filename = payload.get("filename")
    folder_type = payload.get("folder_type", "processed")
    
    if not all([access_token, file_content, filename]):
        raise HTTPException(status_code=400, detail="Missing required parameters")
    
    try:
        import base64
        file_bytes = base64.b64decode(file_content)
        
        service = get_drive_service(access_token)
        
        if folder_type == "processed":
            folder_id = payload.get("processed_folder_id")
        else:
            folder_id = payload.get("unprocessed_folder_id")
        
        if not folder_id:
            raise HTTPException(status_code=400, detail="Folder ID required")
        
        mime_type = "application/octet-stream"
        if filename.lower().endswith('.xlsx'):
            mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif filename.lower().endswith('.xls'):
            mime_type = "application/vnd.ms-excel"
        elif filename.lower().endswith('.pdf'):
            mime_type = "application/pdf"
        
        file_id = upload_file_to_drive(service, file_bytes, filename, folder_id, mime_type)
        
        if file_id:
            return {"success": True, "file_id": file_id, "filename": filename}
        else:
            raise HTTPException(status_code=500, detail="Failed to upload file")
            
    except Exception as e:
        print(f"Drive upload error: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.get("/jobs/{job_id}")
def get_job(job_id: str, x_client_key: str | None = Header(default=None, alias="X-Client-Key")):
    global supabase
    """Get specific job details with retry logic"""
    client_id = get_client_id_from_key(x_client_key)
    max_retries = 3
    for attempt in range(max_retries):
        try:
            res = supabase.table("client_jobs").select("*").eq("id", job_id).eq("client_id", client_id).single().execute()
            job_data = res.data
            
            if job_data:
                created_at = datetime.fromisoformat(job_data["created_at"].replace('Z', '+00:00'))
                elapsed_minutes = (datetime.now(created_at.tzinfo) - created_at).total_seconds() / 60
                
                job_data["elapsed_minutes"] = round(elapsed_minutes, 1)
                
                if job_data["status"] == "pending":
                    job_data["estimated_remaining_minutes"] = job_data.get("estimated_minutes", 10)
                elif job_data["status"] == "processing":
                    estimated_total = job_data.get("estimated_minutes", 10)
                    remaining = max(0, estimated_total - elapsed_minutes)
                    job_data["estimated_remaining_minutes"] = round(remaining, 1)
                else:
                    job_data["estimated_remaining_minutes"] = 0
                
                if "job_data" in job_data:
                    del job_data["job_data"]
            
            return job_data
        except Exception as e:
            print(f"ERROR: Error fetching job {job_id} (attempt {attempt + 1}/{max_retries}): {e}")
            traceback.print_exc()
            try:
                supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
            except Exception as reinit_err:
                print(f"Supabase re-init failed in /jobs/{job_id}: {reinit_err}")
                traceback.print_exc()
            if attempt < max_retries - 1:
                time.sleep(1)
            else:
                print(f"Failed to fetch job {job_id} after {max_retries} attempts")
                raise HTTPException(status_code=500, detail="Database connection failed")

@app.get("/jobs/{job_id}/status")
def get_job_status(job_id: str, x_client_key: str | None = Header(default=None, alias="X-Client-Key")):
    """Get job status for polling - lightweight endpoint"""
    client_id = get_client_id_from_key(x_client_key)
    try:
        res = supabase.table("client_jobs").select("id, status, progress_percent, current_step, created_at, completed_at").eq("id", job_id).eq("client_id", client_id).single().execute()
        job = res.data
        
        if job:
            created_at = datetime.fromisoformat(job["created_at"].replace('Z', '+00:00'))
            elapsed_minutes = (datetime.now(created_at.tzinfo) - created_at).total_seconds() / 60
            
            return {
                "job_id": job["id"],
                "status": job["status"],
                "progress_percent": job["progress_percent"],
                "current_step": job["current_step"],
                "elapsed_minutes": round(elapsed_minutes, 1),
                "completed_at": job.get("completed_at")
            }
        else:
            raise HTTPException(status_code=404, detail="Job not found")
    except Exception as e:
        print(f"ERROR: Error fetching job status {job_id}: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Database connection failed")

@app.delete("/jobs/{job_id}")
def cancel_job(job_id: str, x_client_key: str | None = Header(default=None, alias="X-Client-Key")):
    """Cancel a pending job"""
    client_id = get_client_id_from_key(x_client_key)
    try:
        res = supabase.table("client_jobs").select("status").eq("id", job_id).eq("client_id", client_id).single().execute()
        job_status = res.data.get("status") if res.data else None
    except Exception as e:
        print(f"Error fetching job status for cancellation {job_id}: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Failed to fetch job status for cancellation")

    if job_status in ["pending", "processing"]:
        try:
            supabase.table("client_jobs").update({"status": "cancelled", "current_step": "Job cancelled by user", "completed_at": datetime.now().isoformat()}).eq("id", job_id).eq("client_id", client_id).execute()
            print(f"DEBUG: Job {job_id} cancelled by user.")
            return {"ok": True, "message": f"Job {job_id} cancelled successfully."}
        except Exception as e:
            print(f"Error cancelling job {job_id}: {e}")
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Failed to cancel job {job_id}")
    else:
        print(f"WARN: Attempted to cancel job {job_id} which is in status: {job_status}. Only pending/processing jobs can be cancelled.")
        raise HTTPException(status_code=400, detail=f"Cannot cancel job with status '{job_status}'. Only pending or processing jobs can be cancelled.")

@app.post("/jobs/cleanup")
def cleanup_old_jobs(x_client_key: str | None = Header(default=None, alias="X-Client-Key")):
    """Clean up old completed/failed jobs to free up database space"""
    client_id = get_client_id_from_key(x_client_key)
    
    cutoff_date = datetime.now() - timedelta(days=7)
    
    try:
        res = supabase.table("client_jobs").delete().eq("client_id", client_id).in_("status", ["completed", "failed", "cancelled"]).lt("created_at", cutoff_date.isoformat()).execute()
        
        deleted_count = len(res.data) if res.data else 0
        print(f"DEBUG: Cleaned up {deleted_count} old jobs for client {client_id}")
        
        return {"ok": True, "deleted_count": deleted_count, "message": f"Cleaned up {deleted_count} old jobs"}
    except Exception as e:
        print(f"ERROR: Error cleaning up jobs for client {client_id}: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Failed to cleanup jobs")

@app.get("/jobs/stats")
def get_job_stats(x_client_key: str | None = Header(default=None, alias="X-Client-Key")):
    """Get job statistics for the client"""
    client_id = get_client_id_from_key(x_client_key)
    
    try:
        res = supabase.table("client_jobs").select("status").eq("client_id", client_id).execute()
        jobs = res.data or []
        
        stats = {
            "total": len(jobs),
            "pending": len([j for j in jobs if j["status"] == "pending"]),
            "processing": len([j for j in jobs if j["status"] == "processing"]),
            "completed": len([j for j in jobs if j["status"] == "completed"]),
            "failed": len([j for j in jobs if j["status"] == "failed"]),
            "cancelled": len([j for j in jobs if j["status"] == "cancelled"])
        }
        
        return stats
    except Exception as e:
        print(f"ERROR: Error getting job stats for client {client_id}: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Failed to get job statistics")

# Main entry point for Render deployment
if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)