"""
Excel Processing Service
"""
import io
import time
import gc
import traceback
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from services.gemini_service import get_embedding, detect_questions_in_batch, generate_tailored_answer
from services.supabase_service import search_supabase, pick_best_match
from utils.helpers import format_ai_answer


def _extract_row_data(ws, row_num: int) -> list:
    """
    Extract all cell values from a row as a list.
    Equivalent to: XLSX.utils.sheet_to_json(worksheet, { header: 1 })
    Returns array of cell values for a single row.
    """
    max_col = ws.max_column or 1
    row_values = []
    
    for c_idx in range(1, max_col + 1):
        cell_value = ws.cell(row=row_num, column=c_idx).value
        if cell_value is not None:
            row_values.append(str(cell_value).strip())
        else:
            row_values.append("")
    
    return row_values


def find_first_empty_data_column(ws):
    """
    Finds the first entirely empty column after the rightmost column that contains
    a header in the first row OR any data in any subsequent row.
    Ensures it's not column A (index 1).
    """
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1 
    
    rightmost_filled_col = 1
    # Check row 1 for headers
    for col in range(1, max_col + 1):
        header_cell_value = ws.cell(row=1, column=col).value
        if header_cell_value is not None and str(header_cell_value).strip() != '':
            rightmost_filled_col = max(rightmost_filled_col, col)

    # Check data rows (from row 2 onwards) for content
    for col in range(1, max_col + 1):
        for row in range(2, max_row + 1): 
            data_cell_value = ws.cell(row=row, column=col).value
            if data_cell_value is not None and str(data_cell_value).strip() != '':
                rightmost_filled_col = max(rightmost_filled_col, col)
                break 

    # Start search for empty column from the column immediately after the rightmost filled column.
    # Explicitly ensure AI answers don't go into column A.
    candidate_start_col = max(rightmost_filled_col + 1, 2) 

    current_candidate_col = candidate_start_col
    
    # Add a safety break to prevent infinite loops on malformed or extremely dense sheets
    search_limit = max_col + 50 # Search up to 50 columns beyond current max_col
    while current_candidate_col <= search_limit:
        is_column_empty = True
        for r in range(1, max_row + 1): # Check all rows in the candidate column
            if ws.cell(row=r, column=current_candidate_col).value is not None and \
               str(ws.cell(row=r, column=current_candidate_col).value).strip() != '':
                is_column_empty = False
                break
        
        if is_column_empty:
            print(f"DEBUG: Identified column {get_column_letter(current_candidate_col)} as the first empty data column.")
            return current_candidate_col
        
        current_candidate_col += 1
        
    print(f"WARN: Could not find an entirely empty column after checking {search_limit} columns. Returning {current_candidate_col-1} as fallback.")
    return current_candidate_col -1 # Return the last checked column + 1 as a fallback, or if no empty found, the next


def process_detected_questions_batch(detected_questions: list, client_id: str, rfp_id: str, min_valid_row: int = 2, max_valid_row: int = None) -> dict:
    """
    Process detected questions to generate answers for SPECIFIC rows.
    
    INPUT:  [{"rowNumber": 5, "question": "What is your approach?"}, ...]
    OUTPUT: {5: {"question": "What...", "answer": "Our approach...", "review_status": "..."}}
    
    The returned dictionary maps row numbers directly - no searching needed!
    Row 5 question → Row 5 answer (exact match)
    
    VALIDATION: Ensures answers are never placed in:
    - Row 1 (header row)
    - Rows outside the valid data range (< min_valid_row or > max_valid_row)
    """
    results = {}
    
    for item in detected_questions:
        row_num = item.get("rowNumber", 0)
        question_text = item.get("question", "").strip()
        
        # CRITICAL VALIDATION: Ensure row number is valid
        if not question_text or row_num == 0:
            print(f"WARN: Skipping item with empty question or row 0")
            continue
        
        # NEVER place answers in row 1 (header row)
        if row_num < min_valid_row:
            print(f"ERROR: AI returned invalid row number {row_num} (below min {min_valid_row}). Skipping to prevent header corruption.")
            continue
        
        # Check if row exceeds max valid row
        if max_valid_row and row_num > max_valid_row:
            print(f"ERROR: AI returned invalid row number {row_num} (above max {max_valid_row}). Skipping.")
            continue
        
        try:
            print(f"\n--- Processing Row {row_num} ---")
            print(f"Question: '{question_text[:80]}...'")
            
            # Generate embedding for the question
            emb = get_embedding(question_text)
            if not emb:
                print(f"ERROR: Failed to generate embedding for row {row_num}")
                results[row_num] = {
                    "question": question_text,
                    "answer": "Error: Could not generate embedding",
                    "review_status": "Need review"
                }
                continue
            
            # Search Supabase for matching answers
            matches = search_supabase(emb, client_id, rfp_id)
            
            final_answer = "Not found, needs review."
            review_status = "Need review"
            
            if matches and len(matches) > 0:
                print(f"Found {len(matches)} matches in knowledge base")
                best_match = pick_best_match(matches)
                
                if best_match:
                    similarity = best_match.get("similarity", 0)
                    print(f"Best match similarity: {similarity:.3f}")
                    
                    if similarity >= 0.9:
                        print(f"HIGH CONFIDENCE: Using exact match (similarity >= 0.9)")
                        final_answer = best_match["answer"]
                        review_status = "Approved"
                    else:
                        print(f"LOW CONFIDENCE: Generating tailored answer from {len(matches)} matches")
                        final_answer = generate_tailored_answer(question_text, matches)
                        review_status = "Need review - AI Generated"
                else:
                    print(f"WARN: Matches found but pick_best_match returned None")
            else:
                print(f"No matches found in knowledge base for this question")
            
            formatted_answer = format_ai_answer(final_answer)
            results[row_num] = {
                "question": question_text,
                "answer": formatted_answer,
                "review_status": review_status
            }
            print(f"Result: {review_status}")
            
        except Exception as e:
            print(f"ERROR: Exception processing question for row {row_num}: {e}")
            traceback.print_exc()
            results[row_num] = {
                "question": question_text,
                "answer": "Error processing question",
                "review_status": "Need review"
            }
            continue
    
    return results


def estimate_minutes_from_chars(file_bytes: bytes, job_type: str) -> int:
    """Estimate processing time based on total character count across all sheets.
    Falls back to file-size estimation on error.
    """
    try:
        bio = io.BytesIO(file_bytes)
        xls = pd.ExcelFile(bio)
        total_chars = 0
        for sheet in xls.sheet_names:
            try:
                df = pd.read_excel(xls, sheet_name=sheet, header=None, engine="openpyxl")
                if df is None or df.empty:
                    continue
                # Drop fully empty rows/cols, then convert to strings
                df = df.dropna(how='all')
                df = df.dropna(axis=1, how='all')
                if df.empty:
                    continue
                # Convert to strings and sum lengths; use tab as cell sep and newline as row sep
                # to approximate real characters analyzed
                texts = df.astype(str).apply(lambda row: "\t".join(row.values), axis=1)
                sheet_chars = sum(len(s) for s in texts)
                total_chars += sheet_chars
            except Exception as e:
                print(f"Error in estimate_minutes_from_chars for sheet {sheet}: {e}")
                traceback.print_exc()
                continue

        # Base rates per job type (characters per minute)
        if job_type == "process_rfp":
            chars_per_min = 12000  # more conservative to reflect real processing
            min_minutes, max_minutes = 5, 90
        else:  # extract_qa
            chars_per_min = 20000
            min_minutes, max_minutes = 3, 60

        estimated = int(max(min_minutes, min(max_minutes, (total_chars // max(1, chars_per_min)) + 1)))
        # Ensure at least 1 minute if tiny
        return max(1, estimated)
    except Exception as e:
        print(f"Error in estimate_minutes_from_chars: {e}")
        traceback.print_exc()
        # Fallback: use previous file-size based estimator
        from utils.helpers import estimate_processing_time
        return estimate_processing_time(len(file_bytes), job_type)


def process_excel_file_obj(file_obj: io.BytesIO, filename: str, client_id: str, rfp_id: str = None, job_id: str = None, update_progress_callback=None) -> tuple[io.BytesIO, int, int]:
    """
    Process Excel file, add AI answers, and return the processed file as BytesIO,
    along with counts of processed sheets and questions.
    """
    print(f"DEBUG: process_excel_file_obj started for {filename} (Job ID: {job_id})")
    start_time = time.time()
    max_processing_time = 1800  # 30 minutes max processing time
    
    _processed_sheets_count = 0
    _total_questions_processed = 0

    try:
        original_bytes = file_obj.getvalue()

        wb = openpyxl.load_workbook(io.BytesIO(original_bytes))
        xls = pd.ExcelFile(io.BytesIO(original_bytes))

        total_sheets = len(wb.sheetnames)
        
        ai_col = None
        review_col = None
        
        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            try:
                if time.time() - start_time > max_processing_time:
                    raise Exception(f"Processing timeout: exceeded {max_processing_time/60:.1f} minutes")
                
                if update_progress_callback:
                    update_progress_callback(job_id, 15 + int(sheet_idx * 70 / total_sheets / 2),
                                        f"Processing sheet {sheet_idx + 1}/{total_sheets}: {sheet_name}")
                print(f"DEBUG: Processing sheet {sheet_idx + 1}/{total_sheets}: {sheet_name}")
                ws = wb[sheet_name]

                df = pd.read_excel(xls, sheet_name=sheet_name, header=None, engine="openpyxl")
                
                if df is not None:
                    try:
                        df = df.dropna(how='all')
                        df = df.dropna(axis=1, how='all')
                    except Exception as e:
                        print(f"WARN: Error dropping empty rows/cols in sheet {sheet_name}: {e}")
                        traceback.print_exc()

                if df is None or df.empty:
                    print(f"DEBUG: Sheet {sheet_name} is empty after cleanup, skipping.")
                    _processed_sheets_count += 1
                    continue
                    
                # Initialize AI Answer and Review Status columns
                ai_col = find_first_empty_data_column(ws)
                ws.cell(row=1, column=ai_col, value="AI Answers").alignment = Alignment(wrap_text=True)
                review_col = ai_col + 1
                ws.cell(row=1, column=review_col, value="Review Status").alignment = Alignment(wrap_text=True)

                # Set default column widths for new columns
                ws.column_dimensions[get_column_letter(ai_col)].width = 40
                ws.column_dimensions[get_column_letter(review_col)].width = 15

                max_row = ws.max_row or 1
                min_data_row = 2  # Start from row 2 (skip header)
                batch_size = 50  # Process 50 rows per batch with structured output
                total_questions_in_sheet = 0
                
                # Step 1 & 2: Collect all rows with their data
                rows_with_numbers = []
                
                print(f"DEBUG: Scanning rows {min_data_row} to {max_row} in sheet {sheet_name}")
                
                for row_num in range(min_data_row, max_row + 1):
                    # SAFETY CHECK: Never process row 1 (header)
                    if row_num < min_data_row:
                        print(f"CRITICAL: Skipping row {row_num} - below minimum data row!")
                        continue
                    
                    # Check if row has any content
                    row_has_content = False
                    for col in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row=row_num, column=col).value
                        if cell_value is not None and str(cell_value).strip():
                            row_has_content = True
                            break
                    
                    if row_has_content:
                        row_data = _extract_row_data(ws, row_num)
                        rows_with_numbers.append({
                            "rowNumber": row_num,
                            "rowData": row_data
                        })
                
                total_rows_to_process = len(rows_with_numbers)
                print(f"DEBUG: Collected {total_rows_to_process} non-empty rows from sheet {sheet_name}")
                print(f"DEBUG: Row range being processed: {min_data_row} to {max_row}")
                
                # Validate that no row 1 was included
                invalid_rows = [r["rowNumber"] for r in rows_with_numbers if r["rowNumber"] < min_data_row]
                if invalid_rows:
                    print(f"CRITICAL ERROR: Header rows detected in processing batch: {invalid_rows}. Removing them!")
                    rows_with_numbers = [r for r in rows_with_numbers if r["rowNumber"] >= min_data_row]
                
                # Step 3: Process rows in batches to detect questions
                all_detected_questions = []
                for batch_start in range(0, total_rows_to_process, batch_size):
                    batch_end = min(batch_start + batch_size, total_rows_to_process)
                    batch_rows = rows_with_numbers[batch_start:batch_end]
                    
                    # Update progress
                    if update_progress_callback:
                        progress_pct = 15 + int(sheet_idx * 70 / total_sheets / 2) + int(batch_start * 35 / total_rows_to_process / total_sheets)
                        update_progress_callback(job_id, progress_pct,
                                            f"Detecting questions in rows {batch_start + 1}-{batch_end} of {total_rows_to_process} in sheet {sheet_name}")
                    
                    print(f"DEBUG: Detecting questions in batch {batch_start // batch_size + 1}: rows {batch_start + 1}-{batch_end}")
                    
                    # Call detect_questions_in_batch
                    detected = detect_questions_in_batch(batch_rows)
                    
                    # VALIDATE: Check AI returned row numbers are within valid range
                    if detected:
                        detected_row_nums = [q.get("rowNumber", 0) for q in detected]
                        print(f"DEBUG: AI returned row numbers: {detected_row_nums}")
                        
                        # Check for invalid row numbers
                        invalid_detections = [q for q in detected if q.get("rowNumber", 0) < min_data_row or q.get("rowNumber", 0) > max_row]
                        if invalid_detections:
                            print(f"WARNING: AI returned {len(invalid_detections)} questions with invalid row numbers. Filtering them out.")
                            for inv in invalid_detections:
                                print(f"  - Invalid row {inv.get('rowNumber')}: '{inv.get('question', '')[:50]}...'")
                            # Remove invalid detections
                            detected = [q for q in detected if min_data_row <= q.get("rowNumber", 0) <= max_row]
                    
                    all_detected_questions.extend(detected)
                    print(f"DEBUG: Found {len(detected)} VALID questions in this batch")
                
                print(f"DEBUG: Total {len(all_detected_questions)} questions detected in sheet {sheet_name}")
                
                # Final validation of all detected questions
                if all_detected_questions:
                    all_row_nums = [q.get("rowNumber", 0) for q in all_detected_questions]
                    print(f"DEBUG: All detected row numbers: {sorted(set(all_row_nums))}")
                    
                    # Verify no header rows
                    header_violations = [q for q in all_detected_questions if q.get("rowNumber", 0) < min_data_row]
                    if header_violations:
                        print(f"CRITICAL ERROR: Found {len(header_violations)} questions in header rows!")
                        for hv in header_violations:
                            print(f"  - Row {hv.get('rowNumber')}: '{hv.get('question', '')[:50]}...'")
                        # Remove header violations
                        all_detected_questions = [q for q in all_detected_questions if q.get("rowNumber", 0) >= min_data_row]
                        print(f"DEBUG: After removing header violations: {len(all_detected_questions)} questions remain")
                
                # Step 4 & 5: Generate answers for detected questions
                if all_detected_questions:
                    if update_progress_callback:
                        update_progress_callback(job_id, 15 + int((sheet_idx + 0.5) * 70 / total_sheets),
                                            f"Generating answers for {len(all_detected_questions)} questions in sheet {sheet_name}")
                    
                    # Pass validation parameters to ensure row numbers are correct
                    answers_dict = process_detected_questions_batch(
                        all_detected_questions, 
                        client_id, 
                        rfp_id,
                        min_valid_row=min_data_row,  # Row 2 minimum (never row 1)
                        max_valid_row=max_row          # Maximum row in sheet
                    )
                    
                    print(f"DEBUG: Generated {len(answers_dict)} answers for sheet {sheet_name}")
                    
                    # Step 6: Write answers to EXACT rowNumber
                    # Validate and place each answer IN THE EXACT ROW from AI
                    for row_num, answer_data in answers_dict.items():
                        # TRIPLE VALIDATION: Ensure safe answer placement
                        if row_num < min_data_row:
                            print(f"CRITICAL ERROR: Attempted to place answer in row {row_num} (below min {min_data_row}). SKIPPING to prevent header corruption!")
                            continue
                        
                        if row_num > max_row:
                            print(f"ERROR: Attempted to place answer in row {row_num} (above max {max_row}). SKIPPING!")
                            continue
                        
                        if row_num == 1:
                            print(f"CRITICAL ERROR: Attempted to place answer in HEADER ROW (row 1). SKIPPING!")
                            continue
                        
                        # Safe to place answer - DIRECT ROW MAPPING
                        try:
                            answer_text = answer_data.get("answer", "")
                            review_status = answer_data.get("review_status", "")
                            question_text = answer_data.get("question", "")
                            
                            # Place answer in THE SAME ROW where AI found the question
                            ws.cell(row=row_num, column=ai_col, value=answer_text)
                            ws.cell(row=row_num, column=review_col, value=review_status)
                            total_questions_in_sheet += 1
                            
                            print(f"✓ Row {row_num}: Placed answer for '{question_text[:60]}...'")
                        except Exception as place_error:
                            print(f"ERROR: Failed to place answer in row {row_num}: {place_error}")
                            traceback.print_exc()
                            continue
                
                print(f"\n{'='*80}")
                print(f"SHEET SUMMARY: {sheet_name}")
                print(f"  - Total data rows scanned: {total_rows_to_process}")
                print(f"  - Questions detected by AI: {len(all_detected_questions)}")
                print(f"  - Answers successfully placed: {total_questions_in_sheet}")
                print(f"  - Row range: {min_data_row} to {max_row}")
                if total_questions_in_sheet > 0:
                    placed_rows = [r for r, _ in answers_dict.items() if min_data_row <= r <= max_row]
                    print(f"  - Rows with answers: {sorted(placed_rows)}")
                print(f"{'='*80}\n")
                
                _total_questions_processed += total_questions_in_sheet
                
                del df
                gc.collect()
                _processed_sheets_count += 1
                
            except Exception as sheet_error:
                print(f"ERROR: Failed to process sheet {sheet_name}: {sheet_error}")
                traceback.print_exc()
                continue

        if update_progress_callback:
            update_progress_callback(job_id, 90, "Saving processed Excel file...")
        
        # Format and save workbook
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Adjust column widths for new AI Answer and Review Status columns
            if ai_col and ai_col <= ws.max_column:
                ws.column_dimensions[get_column_letter(ai_col)].width = 40
            if review_col and review_col <= ws.max_column:
                ws.column_dimensions[get_column_letter(review_col)].width = 25
            
            # Enable text wrapping for the AI answer column and review status
            for col_to_wrap in [ai_col, review_col]:
                if col_to_wrap and col_to_wrap <= ws.max_column:
                    for row in range(2, ws.max_row + 1): # Start from row 2 for data cells
                        cell = ws.cell(row=row, column=col_to_wrap)
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Save workbook
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        print(f"DEBUG: process_excel_file_obj completed for {filename} in {time.time() - start_time:.1f}s.")
        return output, _processed_sheets_count, _total_questions_processed
        
    except Exception as e:
        processing_time = time.time() - start_time
        error_msg = f"Excel processing failed after {processing_time:.1f}s: {str(e)}"
        print(f"ERROR: process_excel_file_obj error for {filename}: {error_msg}")
        traceback.print_exc()
        raise Exception(error_msg)

